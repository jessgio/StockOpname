import os
import json
import time
from io import BytesIO
from threading import Lock
from flask import Flask, render_template_string, request, jsonify
import gspread
from google.oauth2.service_account import Credentials
import uuid
from datetime import datetime, timedelta
import pytz
from openpyxl import load_workbook

app = Flask(__name__)

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive"
]

def get_spreadsheet():
    creds_raw = os.environ.get("GOOGLE_CREDENTIALS")
    if not creds_raw:
        raise RuntimeError("GOOGLE_CREDENTIALS environment variable is missing!")
    
    creds_dict = json.loads(creds_raw)
    creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
    client = gspread.authorize(creds)
    return client.open("Aeris Beaute - Stock Opname Master Template")

_SPREADSHEET_CACHE = {"wb": None, "expires": 0.0}
_SUMMARY_CACHE = {"by_session": {}, "expires": 0.0}
_LOOKUPS_CACHE = {"payload": None, "expires": 0.0}
_GUDANG_INDEX_CACHE = {"index": None, "expires": 0.0}
_SKU_LOOKUP_CACHE = {"payload": None, "expires": 0.0}
_DUP_INDEX_CACHE = {}
_SESSIONS_CACHE = {"sessions": None, "expires": 0.0}
_HISTORY_CACHE = {}
_DASHBOARD_CACHE = {"by_session": {}}
_CACHE_LOCK = Lock()
_SPREADSHEET_CACHE_TTL = int(os.environ.get("SPREADSHEET_CACHE_SECONDS", "300"))
_SUMMARY_CACHE_TTL = int(os.environ.get("SUMMARY_CACHE_SECONDS", "30"))
_DASHBOARD_CACHE_TTL = int(os.environ.get("DASHBOARD_CACHE_SECONDS", "20"))
_LOOKUPS_CACHE_TTL = int(os.environ.get("LOOKUPS_CACHE_SECONDS", "300"))
_DUP_INDEX_CACHE_TTL = int(os.environ.get("DUP_INDEX_CACHE_SECONDS", "120"))
_HISTORY_CACHE_TTL = int(os.environ.get("HISTORY_CACHE_SECONDS", "60"))
HISTORY_SCAN_ROWS = int(os.environ.get("HISTORY_SCAN_ROWS", "6000"))
HISTORY_MAX_ITEMS = int(os.environ.get("HISTORY_MAX_ITEMS", "300"))

def get_spreadsheet_cached():
    now = time.time()
    with _CACHE_LOCK:
        if _SPREADSHEET_CACHE["wb"] is not None and now < _SPREADSHEET_CACHE["expires"]:
            return _SPREADSHEET_CACHE["wb"]
    wb = get_spreadsheet()
    with _CACHE_LOCK:
        _SPREADSHEET_CACHE["wb"] = wb
        _SPREADSHEET_CACHE["expires"] = now + _SPREADSHEET_CACHE_TTL
    return wb

def invalidate_dup_index_cache(session_id=None):
    with _CACHE_LOCK:
        if session_id:
            _DUP_INDEX_CACHE.pop(session_id, None)
        else:
            _DUP_INDEX_CACHE.clear()

def history_cache_key(session_id, counter_name):
    return f"{session_id}|{counter_name}"

def invalidate_history_cache(counter_name=None, session_id=None):
    with _CACHE_LOCK:
        if session_id and counter_name:
            _HISTORY_CACHE.pop(history_cache_key(session_id, counter_name), None)
        elif session_id:
            prefix = f"{session_id}|"
            for key in list(_HISTORY_CACHE.keys()):
                if key.startswith(prefix):
                    _HISTORY_CACHE.pop(key, None)
        elif counter_name:
            for key in list(_HISTORY_CACHE.keys()):
                if key.endswith(f"|{counter_name}"):
                    _HISTORY_CACHE.pop(key, None)
        else:
            _HISTORY_CACHE.clear()

def invalidate_summary_cache(session_id=None):
    with _CACHE_LOCK:
        if session_id:
            entry = _SUMMARY_CACHE.get("by_session", {})
            entry.pop(session_id, None)
        else:
            _SUMMARY_CACHE["by_session"] = {}
            _SUMMARY_CACHE["expires"] = 0.0

def invalidate_dashboard_cache(session_id=None):
    with _CACHE_LOCK:
        by_session = _DASHBOARD_CACHE.setdefault("by_session", {})
        if session_id:
            by_session.pop(session_id, None)
        else:
            _DASHBOARD_CACHE["by_session"] = {}

def invalidate_count_caches(counter_name=None, session_id=None, invalidate_dup=True):
    if invalidate_dup:
        invalidate_dup_index_cache(session_id)
    invalidate_history_cache(counter_name, session_id)
    invalidate_summary_cache(session_id)
    invalidate_dashboard_cache(session_id)

def invalidate_gudang_index_cache():
    with _CACHE_LOCK:
        _GUDANG_INDEX_CACHE["index"] = None
        _GUDANG_INDEX_CACHE["expires"] = 0.0

def invalidate_lookups_cache():
    with _CACHE_LOCK:
        _LOOKUPS_CACHE["payload"] = None
        _LOOKUPS_CACHE["expires"] = 0.0

def update_dup_index_entry(session_id, counter_name, loc_string, sku_code, count):
    """Keep duplicate index warm after append without re-reading the sheet."""
    with _CACHE_LOCK:
        entry = _DUP_INDEX_CACHE.get(session_id)
        if not entry or entry.get("index") is None:
            return
        entry["index"][(counter_name, loc_string, sku_code)] = count

def read_raw_counts_for_summary(sheet, session_id=None):
    """Fetch location, SKU, qty (F–H) and Session ID (K) in one request; filter by session."""
    try:
        batch = sheet.batch_get(["F2:H", "K2:K"])
        values = batch[0] if batch else []
        session_col = batch[1] if len(batch) > 1 else []
    except Exception:
        try:
            values = sheet.get("F2:H")
            session_col = sheet.get("K2:K")
        except Exception:
            return []
    records = []
    for i, row in enumerate(values):
        padded = list(row) + ["", "", ""]
        sess_cell = session_col[i] if i < len(session_col) else []
        row_session = str(sess_cell[0] if sess_cell else "").strip()
        if session_id and row_session != session_id:
            continue
        records.append({
            "Precise Location": padded[0],
            "SKU Code": padded[1],
            "Physical Count": padded[2],
        })
    return records

def _row_dict_from_padded(padded):
    return {
        "Log ID": padded[0],
        "Counter Name": padded[1],
        "Counter Team": "",
        "Precise Location": padded[5],
        "SKU Code": padded[6],
        "Physical Count": padded[7],
        "Timestamp": padded[8],
        "Notes": padded[9],
        "Session ID": padded[10] if len(padded) > 10 else "",
    }

def read_raw_count_rows(sheet, start_row=2, end_row=None):
    """Fetch Raw Counts rows (A–K) for a row range."""
    if end_row is None:
        end_row = sheet.row_count
    if end_row < start_row:
        return []
    try:
        values = sheet.get(f"A{start_row}:K{end_row}")
    except Exception:
        return []
    rows = []
    for row in values:
        padded = list(row) + [""] * 11
        rows.append(_row_dict_from_padded(padded))
    return rows

def get_dup_index(sheet, session_id, force_refresh=False):
    """Map (counter, location, sku) -> qty for one session (columns B, F–H, K)."""
    now = time.time()
    if not force_refresh:
        with _CACHE_LOCK:
            entry = _DUP_INDEX_CACHE.get(session_id)
            if entry and entry.get("index") is not None and now < entry.get("expires", 0):
                return entry["index"]

    index = {}
    try:
        batch = sheet.batch_get(["B2:B", "F2:H", "K2:K"])
        counters_col = batch[0] if batch else []
        detail_col = batch[1] if len(batch) > 1 else []
        session_col = batch[2] if len(batch) > 2 else []
    except Exception:
        try:
            counters_col = sheet.get("B2:B")
            detail_col = sheet.get("F2:H")
            session_col = sheet.get("K2:K")
        except Exception:
            counters_col, detail_col, session_col = [], [], []

    counters_col = counters_col or []
    detail_col = detail_col or []
    session_col = session_col or []
    row_count = max(len(counters_col), len(detail_col), len(session_col))
    for i in range(row_count):
        s_row = session_col[i] if i < len(session_col) else []
        row_session = str(s_row[0] if s_row else "").strip()
        if session_id and row_session != session_id:
            continue
        c_row = counters_col[i] if i < len(counters_col) else []
        d_row = detail_col[i] if i < len(detail_col) else []
        counter = str(c_row[0] if c_row else "").strip()
        d_padded = list(d_row) + ["", "", ""]
        loc = str(d_padded[0]).strip()
        sku = str(d_padded[1]).strip()
        if counter and loc and sku:
            index[(counter, loc, sku)] = d_padded[2]

    with _CACHE_LOCK:
        _DUP_INDEX_CACHE[session_id] = {
            "index": index,
            "expires": now + _DUP_INDEX_CACHE_TTL,
        }
    return index

def _history_item_from_row(row):
    return {
        "id": row.get("Log ID"),
        "location": row.get("Precise Location"),
        "sku": row.get("SKU Code"),
        "count": row.get("Physical Count"),
        "timestamp": row.get("Timestamp"),
        "notes": row.get("Notes") or "",
    }

def fetch_counter_history(sheet, session_id, target_name, counter_lookup, force_refresh=False, full_scan=False):
    """Return recent history for one counter in one session."""
    cache_key = history_cache_key(session_id, target_name)
    now = time.time()
    if not force_refresh and not full_scan:
        with _CACHE_LOCK:
            entry = _HISTORY_CACHE.get(cache_key)
            if entry is not None and now < entry["expires"]:
                return entry["items"], entry.get("truncated", False), True

    if full_scan:
        rows = read_raw_count_rows(sheet)
        truncated = False
    else:
        row_count = sheet.row_count
        start_row = max(2, row_count - HISTORY_SCAN_ROWS + 1)
        rows = read_raw_count_rows(sheet, start_row=start_row, end_row=row_count)
        truncated = start_row > 2

    matches = [
        _history_item_from_row(row)
        for row in rows
        if row_matches_session(row, session_id)
        and row_matches_counter(row, target_name, counter_lookup)
    ]
    if len(matches) > HISTORY_MAX_ITEMS:
        truncated = True
        matches = matches[-HISTORY_MAX_ITEMS:]

    items = list(reversed(matches))
    if not full_scan:
        with _CACHE_LOCK:
            _HISTORY_CACHE[cache_key] = {
                "items": items,
                "truncated": truncated,
                "expires": now + _HISTORY_CACHE_TTL,
            }
    return items, truncated, False

def build_lookups_payload(force_refresh=False):
    now = time.time()
    if not force_refresh:
        with _CACHE_LOCK:
            cached = _LOOKUPS_CACHE["payload"]
            if cached is not None and now < _LOOKUPS_CACHE["expires"]:
                return cached, True

    wb = get_spreadsheet_cached()
    location_lookup, loc_warnings = get_valid_locations(wb)
    counter_lookup, counter_warnings = get_valid_counters(wb)
    assignments_payload, assignment_warnings = build_assignments_payload(
        wb, counter_lookup, location_lookup
    )
    payload = {
        "location_lookup": location_lookup,
        "counter_lookup": counter_lookup,
        "lookup_warnings": loc_warnings + counter_warnings + assignment_warnings,
        **assignments_payload,
    }
    with _CACHE_LOCK:
        _LOOKUPS_CACHE["payload"] = payload
        _LOOKUPS_CACHE["expires"] = now + _LOOKUPS_CACHE_TTL
    return payload, False

def get_valid_skus(wb):
    """Build SKU lookup from SKU List tab (SKU Code column)."""
    try:
        sku_worksheet = wb.worksheet("SKU List")
        list_of_lists = sku_worksheet.get_all_values()
    except Exception:
        return {}, []

    if not list_of_lists:
        return {}, []

    headers = list_of_lists[0]
    sku_idx = headers.index("SKU Code") if "SKU Code" in headers else 0
    codes = []
    for row in list_of_lists[1:]:
        if len(row) <= sku_idx:
            continue
        sku_code = str(row[sku_idx]).strip()
        if sku_code:
            codes.append(sku_code)
    return build_lookup(codes, "SKU LIST", "any")

def build_sku_payload(force_refresh=False):
    now = time.time()
    if not force_refresh:
        with _CACHE_LOCK:
            cached = _SKU_LOOKUP_CACHE.get("payload")
            if cached is not None and now < _SKU_LOOKUP_CACHE.get("expires", 0):
                return cached, True

    wb = get_spreadsheet_cached()
    sku_lookup, sku_warnings = get_valid_skus(wb)
    payload = {
        "sku_lookup": sku_lookup,
        "sku_codes": sorted(set(sku_lookup.values())),
        "warnings": sku_warnings,
    }
    with _CACHE_LOCK:
        _SKU_LOOKUP_CACHE["payload"] = payload
        _SKU_LOOKUP_CACHE["expires"] = now + _LOOKUPS_CACHE_TTL
    return payload, False

def load_sku_lookup_cached(wb, force_refresh=False):
    payload, _ = build_sku_payload(force_refresh=force_refresh)
    return payload["sku_lookup"]

def row_matches_counter(row, target_name, counter_lookup):
    row_name = get_row_counter_name(row)
    if not row_name or not target_name:
        return False
    if row_name == target_name:
        return True
    return resolve_counter(row_name, counter_lookup) == target_name

def row_matches_session(row, session_id):
    if not session_id:
        return True
    return str(row.get("Session ID") or "").strip() == session_id

def find_duplicate_count(sheet, session_id, counter_name, loc_string, sku_code):
    """O(1) duplicate lookup for this session only."""
    index = get_dup_index(sheet, session_id)
    return index.get((counter_name, loc_string, sku_code))

def build_summary_payload(session_id, force_refresh=False):
    now = time.time()
    if not force_refresh:
        with _CACHE_LOCK:
            by_session = _SUMMARY_CACHE.get("by_session", {})
            entry = by_session.get(session_id)
            if entry is not None and now < entry.get("expires", 0):
                return entry["payload"], True

    wb = get_spreadsheet_cached()
    sheet = wb.worksheet("Raw Counts")
    records = read_raw_counts_for_summary(sheet, session_id=session_id)
    rows, grand_total = aggregate_counts_by_sku(records)
    payload = {
        "rows": rows,
        "grand_total": grand_total,
        "sku_count": len(rows),
    }
    with _CACHE_LOCK:
        if "by_session" not in _SUMMARY_CACHE:
            _SUMMARY_CACHE["by_session"] = {}
        _SUMMARY_CACHE["by_session"][session_id] = {
            "payload": payload,
            "expires": now + _SUMMARY_CACHE_TTL,
        }
    return payload, False

def parse_opname_timestamp(ts_str):
    try:
        s = str(ts_str or "").strip()
        if not s:
            return None
        return datetime.strptime(s, "%d/%m/%Y %H:%M:%S")
    except Exception:
        return None

def read_session_count_rows(sheet, session_id):
    """Fetch counter, location, SKU, qty, timestamp, session for dashboard metrics."""
    try:
        batch = sheet.batch_get(["B2:B", "F2:H", "I2:I", "K2:K"])
        counters_col = batch[0] if batch else []
        detail_col = batch[1] if len(batch) > 1 else []
        time_col = batch[2] if len(batch) > 2 else []
        session_col = batch[3] if len(batch) > 3 else []
    except Exception:
        return []

    row_count = max(len(counters_col), len(detail_col), len(time_col), len(session_col))
    rows = []
    for i in range(row_count):
        s_row = session_col[i] if i < len(session_col) else []
        row_session = str(s_row[0] if s_row else "").strip()
        if session_id and row_session != session_id:
            continue
        c_row = counters_col[i] if i < len(counters_col) else []
        d_row = detail_col[i] if i < len(detail_col) else []
        t_row = time_col[i] if i < len(time_col) else []
        d_padded = list(d_row) + ["", "", ""]
        rows.append({
            "counter": str(c_row[0] if c_row else "").strip(),
            "location": str(d_padded[0]).strip(),
            "sku": str(d_padded[1]).strip(),
            "qty": parse_physical_count(d_padded[2]),
            "timestamp": str(t_row[0] if t_row else "").strip(),
        })
    return rows

def build_dashboard_payload(session_id, force_refresh=False):
    now = time.time()
    if not force_refresh:
        with _CACHE_LOCK:
            entry = _DASHBOARD_CACHE.get("by_session", {}).get(session_id)
            if entry is not None and now < entry.get("expires", 0):
                return entry["payload"], True

    wb = get_spreadsheet_cached()
    lookups, _ = build_lookups_payload()
    sku_payload, _ = build_sku_payload()
    sessions, _ = build_sessions_payload()
    session_meta = session_by_id(sessions, session_id)

    location_lookup = lookups["location_lookup"]
    sku_lookup = sku_payload["sku_lookup"]
    valid_locations = set(location_lookup.values())
    valid_skus = set(sku_lookup.values())
    locations_total = len(valid_locations)
    skus_total = len(valid_skus)

    sheet = wb.worksheet("Raw Counts")
    rows = read_session_count_rows(sheet, session_id)

    unique_locations = set()
    unique_skus = set()
    unique_counters = set()
    sku_line_counts = {}
    total_qty = 0
    hourly = {}
    jakarta_tz = pytz.timezone("Asia/Jakarta")
    recent_cutoff = datetime.now(jakarta_tz) - timedelta(minutes=15)
    recent_entries = 0

    for row in rows:
        total_qty += row["qty"]
        counter = row["counter"]
        if counter:
            unique_counters.add(counter)

        loc = resolve_location(row["location"], location_lookup) or row["location"]
        if loc and loc in valid_locations:
            unique_locations.add(loc)

        sku = resolve_sku(row["sku"], sku_lookup) or row["sku"]
        if sku and sku in valid_skus:
            unique_skus.add(sku)
            sku_line_counts[sku] = sku_line_counts.get(sku, 0) + 1

        ts = parse_opname_timestamp(row["timestamp"])
        if ts:
            if ts.tzinfo is None:
                ts = jakarta_tz.localize(ts)
            hour_key = ts.strftime("%H:00")
            hourly[hour_key] = hourly.get(hour_key, 0) + 1
            if ts >= recent_cutoff:
                recent_entries += 1

    locations_covered = len(unique_locations)
    skus_covered = len(unique_skus)
    top_skus = sorted(
        [{"sku": k, "lines": v} for k, v in sku_line_counts.items()],
        key=lambda x: (-x["lines"], x["sku"]),
    )[:10]
    hourly_sorted = sorted(
        [{"hour": h, "entries": hourly[h]} for h in hourly],
        key=lambda x: x["hour"],
    )

    payload = {
        "session_id": session_id,
        "session_name": session_meta["name"] if session_meta else session_id,
        "updated_at": datetime.now(jakarta_tz).strftime("%d/%m/%Y %H:%M:%S"),
        "locations": {
            "covered": locations_covered,
            "total": locations_total,
            "pct": round(100 * locations_covered / locations_total, 1) if locations_total else 0,
        },
        "skus": {
            "covered": skus_covered,
            "total": skus_total,
            "pct": round(100 * skus_covered / skus_total, 1) if skus_total else 0,
        },
        "entries": len(rows),
        "total_qty": total_qty,
        "counters_active": len(unique_counters),
        "recent_entries": recent_entries,
        "top_skus": top_skus,
        "hourly": hourly_sorted,
    }
    with _CACHE_LOCK:
        _DASHBOARD_CACHE.setdefault("by_session", {})[session_id] = {
            "payload": payload,
            "expires": now + _DASHBOARD_CACHE_TTL,
        }
    return payload, False

COUNTER_PREFIXES = (
    "counter:", "counter name:", "name:", "nama:", "nama petugas:",
    "petugas:", "id:", "badge:", "id badge:",
)
LOCATION_PREFIXES = (
    "loc:", "location:", "lokasi:", "kode lokasi:", "precise location:",
)

def normalize_scan_text(code, kind="any"):
    """Normalize QR payload or sheet cell. kind: 'counter', 'location', or 'any'."""
    s = str(code or "").strip().replace("\ufeff", "").replace("\u200b", "").replace("\r", "").replace("\n", "")
    if not s:
        return ""
    lower = s.lower()
    if lower.startswith("http://") or lower.startswith("https://") or "://" in s:
        s = s.rstrip("/").split("/")[-1]
    if "?" in s:
        s = s.split("?")[0]
    if "#" in s:
        s = s.split("#")[0]
    s = s.strip()
    if kind == "counter":
        prefixes = COUNTER_PREFIXES
    elif kind == "location":
        prefixes = LOCATION_PREFIXES
    else:
        prefixes = COUNTER_PREFIXES + LOCATION_PREFIXES
    for prefix in prefixes:
        if s.lower().startswith(prefix):
            s = s[len(prefix):].strip()
            break
    return s.strip()

def get_worksheet_by_names(wb, names):
    for name in names:
        try:
            return wb.worksheet(name)
        except Exception:
            continue
    return None

def read_codes_from_sheet(ws, header_aliases, default_col=0):
    """Read codes from the first matching header column, else column A."""
    try:
        rows = ws.get_all_values()
    except Exception:
        return []

    if not rows:
        return []

    header_aliases_lower = {h.lower() for h in header_aliases}
    col_idx = default_col
    data_start = 0

    header_cells = [str(c).strip().lower() for c in rows[0]]
    for i, cell in enumerate(header_cells):
        if cell in header_aliases_lower:
            col_idx = i
            data_start = 1
            break
    else:
        if header_cells and header_cells[default_col] in header_aliases_lower:
            data_start = 1

    codes = []
    for row in rows[data_start:]:
        if col_idx >= len(row):
            continue
        val = str(row[col_idx]).strip()
        if val:
            codes.append(val)
    return codes

def read_codes_from_column_a(ws, header_aliases):
    """Read counter/location names from column A; skip row 1 only if A1 is a header label."""
    try:
        rows = ws.get_all_values()
    except Exception:
        return []

    if not rows:
        return []

    header_aliases_lower = {h.lower() for h in header_aliases}
    start = 0
    if rows[0]:
        first = str(rows[0][0]).strip().lower()
        if first in header_aliases_lower:
            start = 1

    codes = []
    for row in rows[start:]:
        if not row:
            continue
        val = str(row[0]).strip()
        if val:
            codes.append(val)
    return codes

def build_lookup(codes, sheet_label="", normalize_kind="any"):
    """Map normalized lowercase key -> canonical value. Returns (lookup, warning_messages)."""
    lookup = {}
    warnings = []
    for raw in codes:
        canonical = str(raw).strip()
        if not canonical:
            continue
        key = normalize_scan_text(canonical, normalize_kind).lower()
        if not key:
            continue
        if key in lookup and lookup[key] != canonical:
            prefix = f"{sheet_label}: " if sheet_label else ""
            warnings.append(
                f"{prefix}'{lookup[key]}' bentrok dengan '{canonical}' (kode sama setelah normalisasi)"
            )
        lookup[key] = canonical
        plain = " ".join(canonical.lower().split())
        if plain and plain != key:
            if plain in lookup and lookup[plain] != canonical:
                prefix = f"{sheet_label}: " if sheet_label else ""
                warnings.append(
                    f"{prefix}'{lookup[plain]}' bentrok dengan '{canonical}' (varian huruf sama)"
                )
            lookup[plain] = canonical
    return lookup, warnings

def get_valid_locations(wb):
    ws = get_worksheet_by_names(wb, ("LOCATIONS", "Locations", "Location"))
    if not ws:
        return {}, []
    location_aliases = (
        "location", "lokasi", "precise location", "locations", "kode lokasi", "code", "kode",
    )
    codes = read_codes_from_column_a(ws, location_aliases)
    if not codes:
        codes = read_codes_from_sheet(ws, location_aliases)
    return build_lookup(codes, "LOCATIONS", "location")

def resolve_location(code, location_lookup):
    if not code or not location_lookup:
        return None
    trimmed = str(code).strip()
    if trimmed in location_lookup.values():
        return trimmed
    key = normalize_scan_text(trimmed, "location").lower()
    if key and key in location_lookup:
        return location_lookup[key]
    plain = " ".join(trimmed.lower().split())
    if plain and plain in location_lookup:
        return location_lookup[plain]
    return location_lookup.get(key)

def get_valid_counters(wb):
    ws = get_worksheet_by_names(wb, ("COUNTERS", "Counters", "Counter"))
    if not ws:
        return {}, []
    counter_aliases = (
        "counter", "counter name", "nama", "nama petugas", "petugas",
        "name", "counters", "id badge", "badge", "kode",
    )
    # COUNTERS names live in column A; header-in-column-B must not empty the list.
    codes = read_codes_from_column_a(ws, counter_aliases)
    if not codes:
        codes = read_codes_from_sheet(ws, counter_aliases)
    return build_lookup(codes, "COUNTERS", "counter")

def resolve_counter(name, counter_lookup):
    if not name or not counter_lookup:
        return None
    trimmed = str(name).strip()
    if trimmed in counter_lookup.values():
        return trimmed
    key = normalize_scan_text(trimmed, "counter").lower()
    return counter_lookup.get(key)

ASSIGNMENTS_TAB_NAMES = (
    "PETUGAS ASSIGNMENTS",
    "Petugas Assignments",
    "ASSIGNMENTS",
    "Location Assignments",
    "Penugasan Lokasi",
)

def load_petugas_assignments(wb):
    ws = get_worksheet_by_names(wb, ASSIGNMENTS_TAB_NAMES)
    if not ws:
        return []
    try:
        rows = ws.get_all_values()
    except Exception:
        return []
    if not rows:
        return []
    headers = [str(c).strip().lower() for c in rows[0]]
    session_idx = _header_column_index(
        headers, ("session id", "session_id", "id sesi", "sesi id")
    )
    counter_idx = _header_column_index(
        headers,
        (
            "petugas", "counter", "counter name", "nama", "nama petugas",
            "name", "id badge", "badge",
        ),
    )
    location_idx = _header_column_index(
        headers,
        ("location", "lokasi", "precise location", "kode lokasi", "zone", "lokasi ditugaskan"),
    )
    if session_idx is None:
        session_idx = 0
    if counter_idx is None:
        counter_idx = 1
    if location_idx is None:
        location_idx = 2
    out = []
    for row in rows[1:]:
        padded = list(row) + ["", "", ""]
        session_id = str(padded[session_idx]).strip()
        counter = str(padded[counter_idx]).strip()
        location = str(padded[location_idx]).strip()
        if session_id and counter and location:
            out.append({
                "session_id": session_id,
                "counter": counter,
                "location": location,
            })
    return out

def build_assignments_payload(wb, counter_lookup, location_lookup):
    """Per-session petugas→locations map. Sessions with no rows are not enforced."""
    rows = load_petugas_assignments(wb)
    by_session = {}
    enforced_session_ids = set()
    warnings = []
    for item in rows:
        session_id = item["session_id"]
        raw_counter = item["counter"]
        raw_location = item["location"]
        counter = resolve_counter(raw_counter, counter_lookup)
        location = resolve_location(raw_location, location_lookup)
        if raw_counter and counter_lookup and not counter:
            warnings.append(
                f"PETUGAS ASSIGNMENTS: petugas tidak di COUNTERS — {raw_counter!r} (sesi {session_id})"
            )
            continue
        if raw_location and location_lookup and not location:
            warnings.append(
                f"PETUGAS ASSIGNMENTS: lokasi tidak di LOCATIONS — {raw_location!r} (sesi {session_id})"
            )
            continue
        if not counter or not location:
            continue
        enforced_session_ids.add(session_id)
        counter_key = counter.lower()
        by_session.setdefault(session_id, {}).setdefault(counter_key, set()).add(location)
    assignments = {
        sid: {ck: sorted(locs) for ck, locs in counters.items()}
        for sid, counters in by_session.items()
    }
    return (
        {
            "assignments": assignments,
            "enforced_session_ids": sorted(enforced_session_ids),
        },
        warnings,
    )

def ensure_petugas_assignments_tab(wb):
    ws = get_worksheet_by_names(wb, ASSIGNMENTS_TAB_NAMES)
    if ws:
        return ws
    ws = wb.add_worksheet(title="PETUGAS ASSIGNMENTS", rows=500, cols=3)
    ws.append_row(["Session ID", "Petugas", "Location"])
    invalidate_lookups_cache()
    return ws

def save_petugas_assignments(assignments):
    wb = get_spreadsheet_cached()
    lookups, _ = build_lookups_payload(force_refresh=True)
    counter_lookup = lookups.get("counter_lookup") or {}
    location_lookup = lookups.get("location_lookup") or {}
    body = [["Session ID", "Petugas", "Location"]]
    warnings = []
    seen = set()
    for item in assignments:
        if not isinstance(item, dict):
            continue
        session_id = str(item.get("session_id", "")).strip()
        raw_counter = str(item.get("counter", "")).strip()
        raw_location = str(item.get("location", "")).strip()
        if not session_id and not raw_counter and not raw_location:
            continue
        if not session_id:
            warnings.append(f"Baris tanpa Session ID diabaikan (petugas {raw_counter!r}).")
            continue
        if not require_valid_session_id(session_id):
            warnings.append(f"Sesi tidak valid diabaikan: {session_id!r}.")
            continue
        if not raw_counter or not raw_location:
            warnings.append(
                f"Baris tidak lengkap diabaikan (sesi {session_id}, petugas {raw_counter!r})."
            )
            continue
        counter = resolve_counter(raw_counter, counter_lookup)
        location = resolve_location(raw_location, location_lookup)
        if not counter:
            warnings.append(f"Petugas tidak di COUNTERS: {raw_counter!r}.")
            continue
        if not location:
            warnings.append(f"Lokasi tidak di LOCATIONS: {raw_location!r}.")
            continue
        key = (session_id, counter.lower(), location.lower())
        if key in seen:
            continue
        seen.add(key)
        body.append([session_id, counter, location])
    ws = ensure_petugas_assignments_tab(wb)
    ws.clear()
    ws.update(body, range_name="A1")
    invalidate_lookups_cache()
    return len(body) - 1, warnings

def build_petugas_assignments_admin_payload(force_refresh=False):
    wb = get_spreadsheet_cached()
    lookups, from_cache = build_lookups_payload(force_refresh=force_refresh)
    counter_lookup = lookups.get("counter_lookup") or {}
    location_lookup = lookups.get("location_lookup") or {}
    rows = load_petugas_assignments(wb)
    assignments = []
    for item in rows:
        counter = resolve_counter(item["counter"], counter_lookup) or item["counter"]
        location = resolve_location(item["location"], location_lookup) or item["location"]
        assignments.append({
            "session_id": item["session_id"],
            "counter": counter,
            "location": location,
            "valid": bool(
                resolve_counter(item["counter"], counter_lookup)
                and resolve_location(item["location"], location_lookup)
            ),
        })
    sessions, _ = build_sessions_payload(force_refresh=force_refresh)
    assignment_warnings = [
        w for w in (lookups.get("lookup_warnings") or [])
        if "PETUGAS ASSIGNMENTS" in w
    ]
    counters_without = {}
    enforced = set(lookups.get("enforced_session_ids") or [])
    by_session = lookups.get("assignments") or {}
    for sid in enforced:
        session_counters = set(counter_lookup.values())
        assigned = set((by_session.get(sid) or {}).keys())
        missing = sorted(
            c for c in session_counters
            if c.lower() not in assigned
        )
        if missing:
            counters_without[sid] = missing
    return {
        "assignments": assignments,
        "sessions": sessions,
        "counters": sorted(set(counter_lookup.values())),
        "locations": sorted(set(location_lookup.values())),
        "enforced_session_ids": sorted(enforced),
        "counters_without_assignments": counters_without,
        "warnings": assignment_warnings,
    }, from_cache

def session_has_assignment_enforcement(session_id, enforced_session_ids):
    if not session_id or not enforced_session_ids:
        return False
    return session_id in enforced_session_ids

def validate_petugas_location_assignment(
    session_id, counter_name, loc_string, assignments, enforced_session_ids, counter_lookup
):
    if not session_has_assignment_enforcement(session_id, enforced_session_ids):
        return True, None
    canonical = resolve_counter(counter_name, counter_lookup) or counter_name
    counter_key = canonical.lower()
    session_map = assignments.get(session_id) or {}
    allowed = session_map.get(counter_key) or []
    if not allowed:
        return False, (
            f"Petugas {canonical} belum ditugaskan lokasi untuk sesi ini. "
            "Hubungi admin untuk mengisi tab PETUGAS ASSIGNMENTS."
        )
    if loc_string in allowed:
        return True, None
    preview = ", ".join(allowed[:8])
    if len(allowed) > 8:
        preview += f", … (+{len(allowed) - 8} lokasi)"
    return False, (
        f"Lokasi {loc_string} tidak ditugaskan untuk {canonical}. "
        f"Lokasi Anda: {preview}."
    )

def _header_column_index(header_row, aliases):
    for i, cell in enumerate(header_row):
        if str(cell).strip().lower() in aliases:
            return i
    return None

_SESSION_ARCHIVED = frozenset({"archived", "inactive", "closed", "ended"})

def load_sessions_from_sheet(wb):
    ws = get_worksheet_by_names(wb, ("SESSIONS", "Sessions", "Session"))
    if not ws:
        return []
    try:
        rows = ws.get_all_values()
    except Exception:
        return []
    if not rows:
        return []

    headers = [str(c).strip().lower() for c in rows[0]]
    name_idx = _header_column_index(headers, ("session name",))
    id_idx = _header_column_index(headers, ("session id",))
    status_idx = _header_column_index(headers, ("status",))
    if name_idx is None:
        name_idx = 0
    if id_idx is None:
        id_idx = 1

    sessions = []
    for row in rows[1:]:
        padded = list(row) + [""] * 4
        name = str(padded[name_idx]).strip()
        sid = str(padded[id_idx]).strip()
        if not name or not sid:
            continue
        status = str(padded[status_idx]).strip().lower() if status_idx is not None else ""
        if status in _SESSION_ARCHIVED:
            continue
        sessions.append({"id": sid, "name": name, "status": status or "active"})
    return sessions

def build_sessions_payload(force_refresh=False):
    now = time.time()
    if not force_refresh:
        with _CACHE_LOCK:
            cached = _SESSIONS_CACHE.get("sessions")
            if cached is not None and now < _SESSIONS_CACHE.get("expires", 0):
                return cached, True
    wb = get_spreadsheet_cached()
    sessions = load_sessions_from_sheet(wb)
    with _CACHE_LOCK:
        _SESSIONS_CACHE["sessions"] = sessions
        _SESSIONS_CACHE["expires"] = now + _LOOKUPS_CACHE_TTL
    return sessions, False

def session_by_id(sessions, session_id):
    sid = str(session_id or "").strip()
    if not sid:
        return None
    for s in sessions:
        if s["id"] == sid:
            return s
    return None

def require_valid_session_id(session_id):
    sessions, _ = build_sessions_payload()
    return session_by_id(sessions, session_id)

def resolve_sku(code, sku_lookup):
    """Match scanned or typed SKU to a canonical code in the SKU lookup."""
    if not code or not sku_lookup:
        return None
    trimmed = str(code).strip()
    if trimmed in sku_lookup.values():
        return trimmed
    key = normalize_scan_text(trimmed).lower()
    if key and key in sku_lookup:
        return sku_lookup[key]
    plain = " ".join(trimmed.lower().split())
    if plain and plain in sku_lookup:
        return sku_lookup[plain]
    return sku_lookup.get(key)

def get_row_counter_name(row):
    """Read counter from new or legacy column header."""
    return str(row.get("Counter Name") or row.get("Counter Team") or "").strip()

def parse_physical_count(value):
    try:
        return int(value)
    except (ValueError, TypeError):
        return 0

def aggregate_counts_by_sku(records):
    """Sum Physical Count per SKU Code from Raw Counts rows."""
    by_sku = {}
    for row in records:
        sku = str(row.get("SKU Code", "")).strip()
        if not sku:
            continue
        count = parse_physical_count(row.get("Physical Count"))
        if sku not in by_sku:
            by_sku[sku] = {"total": 0, "locations": set()}
        by_sku[sku]["total"] += count
        loc = str(row.get("Precise Location", "")).strip()
        if loc:
            by_sku[sku]["locations"].add(loc)

    rows = [
        {
            "sku": sku,
            "total": data["total"],
            "location_count": len(data["locations"]),
        }
        for sku, data in sorted(by_sku.items())
    ]
    grand_total = sum(r["total"] for r in rows)
    return rows, grand_total

GUDANG_LOCATION_TAB_NAMES = ("GUDANG LOCATIONS", "Gudang Locations", "Location Gudang")
STOCK_RECON_HEADERS = ("Gudang", "SKU", "System Qty", "Running Count", "Gap", "Variance %")
_EXCEL_GUDANG_ROW = 5
_EXCEL_SKU_COL = 3
_EXCEL_DATA_START_ROW = 6

def sanitize_worksheet_title(name):
    s = str(name or "").strip()
    for ch in (":", "\\", "/", "?", "*", "[", "]"):
        s = s.replace(ch, "-")
    return (s[:100] if s else "Session")

def parse_excel_quantity(value):
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return int(round(value))
    s = str(value).strip().replace(",", "")
    if not s:
        return 0
    try:
        return int(round(float(s)))
    except (ValueError, TypeError):
        return 0

def normalize_gudang_label(name):
    return " ".join(str(name or "").strip().split())

def ensure_gudang_locations_tab(wb):
    ws = get_worksheet_by_names(wb, GUDANG_LOCATION_TAB_NAMES)
    if ws:
        return ws
    ws = wb.add_worksheet(title="GUDANG LOCATIONS", rows=500, cols=3)
    ws.append_row(["Location", "Gudang", "Notes"])
    invalidate_gudang_index_cache()
    return ws

def load_gudang_location_mappings(wb):
    ws = get_worksheet_by_names(wb, GUDANG_LOCATION_TAB_NAMES)
    if not ws:
        return []
    try:
        rows = ws.get_all_values()
    except Exception:
        return []
    if not rows:
        return []
    headers = [str(c).strip().lower() for c in rows[0]]
    loc_idx = _header_column_index(headers, ("location", "lokasi", "precise location", "kode lokasi", "zone"))
    gudang_idx = _header_column_index(headers, ("gudang", "warehouse", "gudang name"))
    notes_idx = _header_column_index(headers, ("notes", "catatan", "keterangan"))
    if loc_idx is None:
        loc_idx = 0
    if gudang_idx is None:
        gudang_idx = 1
    mappings = []
    for row in rows[1:]:
        padded = list(row) + ["", "", ""]
        location = str(padded[loc_idx]).strip()
        gudang = normalize_gudang_label(padded[gudang_idx])
        notes = str(padded[notes_idx]).strip() if notes_idx is not None else ""
        if location and gudang:
            mappings.append({"location": location, "gudang": gudang, "notes": notes})
    return mappings

def build_gudang_location_index(wb, force_refresh=False):
    now = time.time()
    if not force_refresh:
        with _CACHE_LOCK:
            cached = _GUDANG_INDEX_CACHE.get("index")
            if cached is not None and now < _GUDANG_INDEX_CACHE.get("expires", 0):
                return cached, True
    mappings = load_gudang_location_mappings(wb)
    index = {}
    for item in mappings:
        loc = item["location"]
        gudang = item["gudang"]
        index[loc.lower()] = gudang
        index[normalize_scan_text(loc, "location").lower()] = gudang
        zone = loc.split("-")[0].strip().lower() if "-" in loc else ""
        if zone and zone not in index:
            index[zone] = gudang
    with _CACHE_LOCK:
        _GUDANG_INDEX_CACHE["index"] = index
        _GUDANG_INDEX_CACHE["expires"] = now + _LOOKUPS_CACHE_TTL
    return index, False

def resolve_gudang_for_location(location, gudang_index):
    if not location or not gudang_index:
        return None
    loc = str(location).strip()
    if not loc:
        return None
    if loc.lower() in gudang_index:
        return gudang_index[loc.lower()]
    norm = normalize_scan_text(loc, "location").lower()
    if norm in gudang_index:
        return gudang_index[norm]
    zone = loc.split("-")[0].strip().lower() if "-" in loc else loc.lower()
    return gudang_index.get(zone)

def aggregate_counts_by_gudang_sku(records, gudang_index):
    """Sum physical counts per (gudang, sku) using the location→gudang index."""
    totals = {}
    unmapped = set()
    for row in records:
        sku = str(row.get("SKU Code", "")).strip()
        loc = str(row.get("Precise Location", "")).strip()
        if not sku or not loc:
            continue
        gudang = resolve_gudang_for_location(loc, gudang_index)
        if not gudang:
            unmapped.add(loc)
            continue
        qty = parse_physical_count(row.get("Physical Count"))
        key = (gudang, sku)
        totals[key] = totals.get(key, 0) + qty
    return totals, unmapped

def _is_gudang_stock_column(label):
    """Match warehouse qty columns; skip totals/subtotals (e.g. 'Total Nama Gudang')."""
    low = label.lower()
    if "gudang" not in low:
        return False
    if "total" in low or "jumlah" in low or "subtotal" in low:
        return False
    return True

def _is_excel_summary_sku_row(sku):
    """Skip ERP footer/summary rows (e.g. 'Total Kode Barang')."""
    low = " ".join(str(sku or "").strip().lower().split())
    if not low:
        return True
    if low == "total kode barang":
        return True
    if low.startswith("total") and "kode" in low:
        return True
    return False

def parse_system_stock_excel(file_bytes, sku_lookup=None):
    """Parse ERP export: gudang names on row 5, SKU in column C from row 6."""
    # read_only breaks max_row on some ERP exports; load fully for reliable dimensions.
    wb_xl = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb_xl.active
    gudang_cols = {}
    for col in range(1, (ws.max_column or 0) + 1):
        val = ws.cell(row=_EXCEL_GUDANG_ROW, column=col).value
        if val is None:
            continue
        label = normalize_gudang_label(val)
        if label and _is_gudang_stock_column(label):
            gudang_cols[col] = label
    if not gudang_cols:
        raise ValueError(
            "Tidak menemukan nama gudang di baris 5. "
            'Pastikan ada sel seperti "Gudang Finished Goods", "Gudang Raw", dll.'
        )
    if not sku_lookup:
        wb_xl.close()
        raise ValueError("Daftar SKU tidak tersedia. Periksa tab SKU List di sheet sebelum upload.")
    rows_out = []
    warnings = []
    excluded_unrecognized = 0
    max_row = ws.max_row or _EXCEL_DATA_START_ROW
    for row_idx in range(_EXCEL_DATA_START_ROW, max_row + 1):
        raw_sku = ws.cell(row=row_idx, column=_EXCEL_SKU_COL).value
        if raw_sku is None or str(raw_sku).strip() == "":
            continue
        sku = str(raw_sku).strip()
        if _is_excel_summary_sku_row(sku):
            continue
        resolved = resolve_sku(sku, sku_lookup)
        if not resolved:
            excluded_unrecognized += 1
            warnings.append(f"SKU tidak dikenali (baris {row_idx}): {sku}")
            continue
        sku = resolved
        for col_idx, gudang in gudang_cols.items():
            qty = parse_excel_quantity(ws.cell(row=row_idx, column=col_idx).value)
            if qty == 0:
                continue
            rows_out.append({"gudang": gudang, "sku": sku, "system_qty": qty})
    wb_xl.close()
    if not rows_out:
        raise ValueError(
            "Tidak ada baris stok sistem yang terbaca. "
            "Pastikan SKU ada di tab SKU List dan kuantitas gudang tidak nol."
        )
    return rows_out, warnings, excluded_unrecognized

def stock_variance_pct(system_qty, running_qty):
    gap = running_qty - system_qty
    if system_qty:
        return round((gap / system_qty) * 100, 2)
    if running_qty:
        return 100.0
    return 0.0

def get_session_stock_worksheet(wb, session_id, create=False):
    title = sanitize_worksheet_title(session_id)
    try:
        return wb.worksheet(title)
    except Exception:
        if not create:
            return None
    ws = wb.add_worksheet(title=title, rows=max(len(STOCK_RECON_HEADERS) + 1, 500), cols=len(STOCK_RECON_HEADERS))
    ws.append_row(list(STOCK_RECON_HEADERS))
    return ws

def session_stock_tab_exists(wb, session_id):
    return get_session_stock_worksheet(wb, session_id, create=False) is not None

def write_session_stock_tab(ws, stock_rows, running_by_key):
    """Rewrite session tab with system stock and computed running/gap/variance."""
    body = [list(STOCK_RECON_HEADERS)]
    for item in stock_rows:
        gudang = item["gudang"]
        sku = item["sku"]
        system_qty = int(item.get("system_qty", 0))
        running = int(running_by_key.get((gudang, sku), 0))
        gap = running - system_qty
        variance = stock_variance_pct(system_qty, running)
        body.append([gudang, sku, system_qty, running, gap, f"{variance}%"])
    ws.clear()
    ws.update(body, range_name="A1")
    return len(body) - 1

def refresh_session_stock_tab(session_id):
    """Recalculate running count, gap, and variance on an existing session stock tab."""
    wb = get_spreadsheet_cached()
    ws = get_session_stock_worksheet(wb, session_id, create=False)
    if not ws:
        return {"updated": 0, "skipped": True}
    try:
        data = ws.get_all_values()
    except Exception:
        return {"updated": 0, "error": "read_failed"}
    if len(data) < 2:
        return {"updated": 0}
    gudang_index, _ = build_gudang_location_index(wb)
    sheet = wb.worksheet("Raw Counts")
    records = read_raw_counts_for_summary(sheet, session_id=session_id)
    running_by_key, _ = aggregate_counts_by_gudang_sku(records, gudang_index)
    updates = []
    for i, row in enumerate(data[1:], start=2):
        padded = list(row) + [""] * 6
        gudang = normalize_gudang_label(padded[0])
        sku = str(padded[1]).strip()
        if not gudang or not sku:
            continue
        system_qty = parse_physical_count(padded[2])
        running = int(running_by_key.get((gudang, sku), 0))
        gap = running - system_qty
        variance = stock_variance_pct(system_qty, running)
        updates.append([running, gap, f"{variance}%"])
    if not updates:
        return {"updated": 0}
    ws.update(updates, range_name=f"D2:F{len(updates) + 1}", value_input_option="USER_ENTERED")
    return {"updated": len(updates)}

def import_system_stock_for_session(session_id, file_bytes):
    if not require_valid_session_id(session_id):
        raise ValueError("Sesi tidak valid.")
    wb = get_spreadsheet_cached()
    sku_lookup = load_sku_lookup_cached(wb)
    stock_rows, warnings, excluded_unrecognized = parse_system_stock_excel(
        file_bytes, sku_lookup=sku_lookup
    )
    gudang_index, _ = build_gudang_location_index(wb)
    sheet = wb.worksheet("Raw Counts")
    records = read_raw_counts_for_summary(sheet, session_id=session_id)
    running_by_key, unmapped = aggregate_counts_by_gudang_sku(records, gudang_index)
    ws = get_session_stock_worksheet(wb, session_id, create=True)
    row_count = write_session_stock_tab(ws, stock_rows, running_by_key)
    return {
        "row_count": row_count,
        "tab_title": sanitize_worksheet_title(session_id),
        "warnings": warnings,
        "excluded_unrecognized": excluded_unrecognized,
        "unmapped_locations": sorted(unmapped),
    }

def save_gudang_location_mappings(mappings):
    wb = get_spreadsheet_cached()
    ws = ensure_gudang_locations_tab(wb)
    body = [["Location", "Gudang", "Notes"]]
    for item in mappings:
        loc = str(item.get("location", "")).strip()
        gudang = normalize_gudang_label(item.get("gudang", ""))
        notes = str(item.get("notes", "")).strip()
        if loc and gudang:
            body.append([loc, gudang, notes])
    ws.clear()
    ws.update(body, range_name="A1")
    invalidate_gudang_index_cache()
    return len(body) - 1

def build_gudang_locations_payload(force_refresh=False):
    wb = get_spreadsheet_cached()
    index, from_cache = build_gudang_location_index(wb, force_refresh=force_refresh)
    mappings = load_gudang_location_mappings(wb)
    lookups, _ = build_lookups_payload()
    location_lookup = lookups.get("location_lookup", {})
    unmapped = sorted(
        loc for loc in set(location_lookup.values())
        if not resolve_gudang_for_location(loc, index)
    )
    return {
        "unmapped_locations": unmapped,
        "all_mapped": len(unmapped) == 0,
        "mapping_count": len(mappings),
    }, from_cache

def maybe_refresh_session_stock(session_id):
    if not session_id:
        return
    try:
        wb = get_spreadsheet_cached()
        if session_stock_tab_exists(wb, session_id):
            refresh_session_stock_tab(session_id)
    except Exception:
        pass

SESSION_HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, viewport-fit=cover">
    <title>Start Session — Aeris Opname</title>
    <script src="https://cdn.jsdelivr.net/npm/@tailwindcss/browser@4"></script>
</head>
<body class="bg-zinc-50 font-sans text-zinc-900 antialiased min-h-screen flex flex-col">
    <header class="border-b border-zinc-200 bg-white shadow-sm">
        <div class="max-w-md mx-auto px-4 py-3">
            <h1 class="text-lg font-bold text-zinc-900 tracking-tight">Aeris Beaute</h1>
            <p class="text-xs text-zinc-500">Stock Opname — pilih sesi</p>
        </div>
    </header>
    <main class="flex-1 max-w-md w-full mx-auto px-4 py-8">
        <div id="continueBox" class="hidden mb-6 rounded-xl border border-violet-200 bg-violet-50 px-4 py-3">
            <p class="text-sm text-violet-900">Lanjutkan sesi:</p>
            <p id="continueLabel" class="font-semibold text-violet-950 mt-1"></p>
            <div class="flex gap-2 mt-3">
                <a href="/count" class="flex-1 text-center py-2.5 rounded-lg bg-violet-600 text-white text-sm font-semibold">Ke halaman Count</a>
                <a href="/dashboard" class="px-3 py-2.5 rounded-lg border border-violet-300 text-violet-800 text-sm font-semibold text-center">Dashboard</a>
                <button type="button" onclick="endStoredSession()" class="px-3 py-2.5 rounded-lg border border-zinc-300 text-zinc-700 text-sm font-semibold">End Session</button>
            </div>
        </div>
        <div class="bg-white rounded-xl border border-zinc-200 shadow-sm p-5 space-y-4">
            <div>
                <label for="sessionSelect" class="block text-sm font-semibold text-zinc-700 mb-2">Session Name</label>
                <select id="sessionSelect" class="w-full border border-zinc-200 rounded-lg px-3 py-3 text-sm font-medium bg-white focus:border-violet-500 focus:ring-2 focus:ring-violet-500/20 focus:outline-none">
                    <option value="">Memuat daftar sesi…</option>
                </select>
            </div>
            <button type="button" id="startSessionBtn" onclick="startSession()" disabled
                class="w-full py-3 rounded-lg bg-violet-600 hover:bg-violet-700 disabled:opacity-50 disabled:cursor-not-allowed text-white text-sm font-semibold transition">
                Start Session
            </button>
            <p id="sessionError" class="hidden text-sm text-rose-600"></p>
        </div>
        <p class="text-xs text-zinc-400 text-center mt-6">End Session hanya di perangkat ini — tidak menutup sesi di sheet.</p>
        <p class="text-center mt-3"><a href="/admin/stock" class="text-xs font-semibold text-violet-600 hover:text-violet-800">Admin: stok sistem &amp; penugasan lokasi</a></p>
    </main>
    <script>
        const SESSION_STORAGE_KEY = 'aeris_opname_session';
        let sessionsList = [];

        function loadStoredSession() {
            try {
                const raw = localStorage.getItem(SESSION_STORAGE_KEY);
                return raw ? JSON.parse(raw) : null;
            } catch (e) { return null; }
        }

        function endStoredSession() {
            localStorage.removeItem(SESSION_STORAGE_KEY);
            document.getElementById('continueBox').classList.add('hidden');
        }

        function showContinue(stored) {
            if (!stored || !stored.sessionId) return;
            document.getElementById('continueLabel').textContent = stored.sessionName || stored.sessionId;
            document.getElementById('continueBox').classList.remove('hidden');
        }

        async function loadSessions() {
            const sel = document.getElementById('sessionSelect');
            const err = document.getElementById('sessionError');
            try {
                const res = await fetch('/api/sessions');
                const data = await res.json();
                sessionsList = data.sessions || [];
                sel.innerHTML = '';
                if (!sessionsList.length) {
                    sel.innerHTML = '<option value="">Tidak ada sesi aktif di tab SESSIONS</option>';
                    document.getElementById('startSessionBtn').disabled = true;
                    return;
                }
                sessionsList.forEach(s => {
                    const opt = document.createElement('option');
                    opt.value = s.id;
                    opt.textContent = s.name;
                    sel.appendChild(opt);
                });
                document.getElementById('startSessionBtn').disabled = false;
                const stored = loadStoredSession();
                if (stored && stored.sessionId) {
                    const match = sessionsList.find(s => s.id === stored.sessionId);
                    if (match) sel.value = match.id;
                }
            } catch (e) {
                err.textContent = 'Gagal memuat daftar sesi.';
                err.classList.remove('hidden');
                sel.innerHTML = '<option value="">Error</option>';
            }
        }

        function startSession() {
            const sel = document.getElementById('sessionSelect');
            const err = document.getElementById('sessionError');
            const id = sel.value;
            const session = sessionsList.find(s => s.id === id);
            if (!session) {
                err.textContent = 'Pilih sesi terlebih dahulu.';
                err.classList.remove('hidden');
                return;
            }
            localStorage.setItem(SESSION_STORAGE_KEY, JSON.stringify({
                sessionId: session.id,
                sessionName: session.name,
                startedAt: Date.now(),
            }));
            window.location.href = '/count';
        }

        const stored = loadStoredSession();
        showContinue(stored);
        loadSessions();
    </script>
</body>
</html>
"""

# --- HTML INTERFACE WITH DUPLICATE PROTECTION & AUTO-RESET ---
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, viewport-fit=cover">
    <title>Aeris Opname 2026</title>
    <style>
        #scanModal { display: none !important; visibility: hidden !important; pointer-events: none !important; }
        #scanModal.is-open { display: block !important; visibility: visible !important; pointer-events: auto !important; }
        #toast { pointer-events: none !important; }
        #step1Card { position: relative; z-index: 1; pointer-events: auto !important; }
        #step1Card input, #step1Card button { pointer-events: auto !important; touch-action: manipulation; }
        .opname-field {
            font-family: inherit;
            font-size: 1rem;
            line-height: 1.5;
            font-weight: 600;
        }
        .opname-field::placeholder { font-weight: 500; color: rgb(251 191 36 / 0.75); }
        .step-card { border-radius: 0.75rem; border: 1px solid #e4e4e7; background: #fff; box-shadow: 0 1px 2px rgb(0 0 0 / 0.05); overflow: hidden; transition: opacity 0.2s, border-color 0.2s; }
        .step-card--active { border-color: #ddd6fe; border-width: 2px; }
        .step-card__head { display: flex; align-items: center; gap: 0.75rem; padding: 0.75rem 1rem; border-bottom: 1px solid #f4f4f5; }
        .step-card__head--active { background: rgb(245 243 255 / 0.9); border-bottom-color: #ede9fe; }
        .step-card__head--idle { background: #fafafa; }
        .step-card__badge { font-size: 0.75rem; font-weight: 600; padding: 0.125rem 0.5rem; border-radius: 0.25rem; }
        .step-card__badge--active { color: #6d28d9; background: #ede9fe; }
        .step-card__badge--idle { color: #71717a; background: #f4f4f5; }
        .step-card__title { font-size: 0.875rem; font-weight: 600; }
        .step-card__title--active { color: #4c1d95; }
        .step-card__title--idle { color: #52525b; }
        .step-locked { opacity: 0.65; }
        .step-locked button, .step-locked select { pointer-events: none; }
        .step-locked input:not([readonly]) { pointer-events: none; }
        #opnameTopChrome {
            position: sticky;
            top: 0;
            z-index: 50;
            background: #fff;
        }
        body.opname-location-locked #opnameTopChrome {
            position: fixed;
            left: 0;
            right: 0;
            top: 0;
            z-index: 50;
            box-shadow: 0 2px 12px rgb(0 0 0 / 0.1);
        }
        body.opname-location-locked main {
            padding-top: var(--opname-chrome-h, 7.25rem);
        }
        .location-sticky-bar {
            background: rgb(236 253 245 / 0.98);
            backdrop-filter: blur(8px);
            border-bottom: 2px solid #34d399;
            box-shadow: 0 2px 8px rgb(16 185 129 / 0.15);
        }
        .location-sticky-bar.hidden { display: none !important; }
        .opname-history-sticky { top: 1rem; }
        body.opname-location-locked .opname-history-sticky {
            top: calc(var(--opname-chrome-h, 7.25rem) + 0.5rem);
        }
        body.opname-location-locked .opname-history-sticky {
            max-height: calc(100vh - var(--opname-chrome-h, 7.25rem) - 2rem);
        }
        body.opname-location-locked #historyContainer {
            max-height: calc(100vh - var(--opname-chrome-h, 7.25rem) - 6rem);
        }
        /* Step 3: allow SKU autocomplete to paint over Step 4+ (step-card defaults to overflow:hidden) */
        #step3Card.step-card {
            overflow: visible;
            position: relative;
        }
        #step3Card.step-card.sku-suggest-open {
            z-index: 45;
        }
        .sku-field-wrap { position: relative; z-index: 1; }
        #skuSuggestions {
            position: absolute;
            left: 0;
            right: 0;
            top: calc(100% + 4px);
            z-index: 50;
            max-height: 12rem;
            overflow-y: auto;
            border-radius: 0.5rem;
            border: 1px solid #e4e4e7;
            background: #fff;
            box-shadow: 0 8px 24px rgb(0 0 0 / 0.12);
        }
        #skuSuggestions.hidden { display: none !important; }
        #skuSuggestions li {
            padding: 0.625rem 0.75rem;
            font-size: 0.875rem;
            font-weight: 600;
            font-family: ui-monospace, monospace;
            color: #5b21b6;
            cursor: pointer;
        }
        #skuSuggestions li:hover, #skuSuggestions li[aria-selected="true"] {
            background: rgb(245 243 255);
        }
    </style>
    <script src="https://cdn.jsdelivr.net/npm/@tailwindcss/browser@4"></script>
</head>
<body class="bg-zinc-50 font-sans text-zinc-900 antialiased min-h-screen flex flex-col">
    <script>
        (function () {
            function unblockPage() {
                document.body.style.overflow = '';
                document.body.style.pointerEvents = '';
                var modal = document.getElementById('scanModal');
                if (modal) {
                    modal.hidden = true;
                    modal.classList.remove('is-open');
                    modal.style.display = 'none';
                }
            }
            if (document.readyState === 'loading') {
                document.addEventListener('DOMContentLoaded', unblockPage);
            } else {
                unblockPage();
            }
            window.addEventListener('pageshow', unblockPage);
        })();
    </script>

    <!-- Toast -->
    <div id="toast" class="fixed top-4 left-4 right-4 z-[60] mx-auto max-w-md translate-y-[-120%] opacity-0 transition-all duration-300 pointer-events-none">
        <div id="toastInner" class="rounded-xl px-4 py-3 text-sm font-medium shadow-lg border"></div>
    </div>

    <!-- Sticky top: brand header + locked location (stays above scrolling form) -->
    <div id="opnameTopChrome">
        <header class="border-b border-zinc-200 shadow-sm">
            <div class="max-w-md lg:max-w-5xl mx-auto px-4 py-3">
                <div class="flex items-center justify-between gap-3">
                    <div>
                        <h1 class="text-lg font-bold text-zinc-900 tracking-tight">Aeris Beaute</h1>
                        <p class="text-xs text-zinc-500">Stock Opname 2026</p>
                    </div>
                    <nav class="flex flex-col items-end gap-1 text-xs font-semibold shrink-0">
                        <div class="flex flex-wrap gap-3 justify-end">
                            <span class="text-violet-700">Count</span>
                            <a href="/dashboard" class="text-zinc-500 hover:text-violet-700">Dashboard</a>
                            <a href="/summary" class="text-zinc-500 hover:text-violet-700">Summary</a>
                        </div>
                        <p id="sessionBadge" class="text-[10px] text-zinc-500 max-w-[10rem] truncate"></p>
                        <button type="button" onclick="endSession()" class="text-[10px] text-rose-600 hover:text-rose-800 font-semibold">End Session</button>
                    </nav>
                </div>
            </div>
        </header>
        <div id="locationStickyBar" class="location-sticky-bar hidden" aria-live="polite">
            <div class="max-w-md lg:max-w-5xl mx-auto px-4 py-3 flex items-center gap-3 min-h-[3rem]">
                <span class="text-xs font-semibold uppercase tracking-wide text-emerald-800 shrink-0">Lokasi</span>
                <span id="locationStickyValue" class="flex-1 min-w-0 text-base sm:text-lg font-bold text-emerald-950 truncate leading-tight"></span>
                <button type="button" id="changeLocationBtn" class="shrink-0 text-xs font-semibold text-emerald-800 bg-emerald-100 hover:bg-emerald-200 px-2.5 py-1.5 rounded-md">Ubah</button>
            </div>
        </div>
    </div>

    <div id="lookupWarnings" class="hidden max-w-md lg:max-w-5xl mx-auto px-4">
        <div class="rounded-lg border border-amber-200 bg-amber-50 px-3 py-2 text-xs text-amber-900 space-y-1"></div>
    </div>

    <main class="flex-1 max-w-md lg:max-w-5xl w-full mx-auto px-4 py-4 pb-28 lg:pb-6 relative z-0">

        <!-- Stepper: Petugas → Lokasi → Produk → Jumlah -->
        <nav class="flex items-center gap-1 sm:gap-2 mb-5" aria-label="Count progress">
            <div id="step1Indicator" class="flex-1 flex items-center gap-1 min-w-0">
                <span id="step1Dot" class="shrink-0 flex h-7 w-7 sm:h-8 sm:w-8 items-center justify-center rounded-full text-xs font-bold border-2 border-violet-600 bg-violet-600 text-white">1</span>
                <span class="text-[10px] sm:text-xs font-medium text-zinc-700 truncate hidden sm:inline">Petugas</span>
            </div>
            <div class="h-px w-2 sm:w-4 bg-zinc-200 shrink-0" aria-hidden="true"></div>
            <div id="step2Indicator" class="flex-1 flex items-center gap-1 min-w-0">
                <span id="step2Dot" class="shrink-0 flex h-7 w-7 sm:h-8 sm:w-8 items-center justify-center rounded-full text-xs font-bold border-2 border-zinc-200 bg-white text-zinc-400">2</span>
                <span class="text-[10px] sm:text-xs font-medium text-zinc-400 truncate hidden sm:inline">Lokasi</span>
            </div>
            <div class="h-px w-2 sm:w-4 bg-zinc-200 shrink-0" aria-hidden="true"></div>
            <div id="step3Indicator" class="flex-1 flex items-center gap-1 min-w-0">
                <span id="step3Dot" class="shrink-0 flex h-7 w-7 sm:h-8 sm:w-8 items-center justify-center rounded-full text-xs font-bold border-2 border-zinc-200 bg-white text-zinc-400">3</span>
                <span class="text-[10px] sm:text-xs font-medium text-zinc-400 truncate hidden sm:inline">Produk</span>
            </div>
            <div class="h-px w-2 sm:w-4 bg-zinc-200 shrink-0" aria-hidden="true"></div>
            <div id="step4Indicator" class="flex-1 flex items-center gap-1 min-w-0">
                <span id="step4Dot" class="shrink-0 flex h-7 w-7 sm:h-8 sm:w-8 items-center justify-center rounded-full text-xs font-bold border-2 border-zinc-200 bg-white text-zinc-400">4</span>
                <span class="text-[10px] sm:text-xs font-medium text-zinc-400 truncate hidden sm:inline">Jumlah</span>
            </div>
        </nav>

        <!-- Mobile tabs -->
        <div class="lg:hidden flex rounded-lg bg-zinc-100 p-1 mb-4" role="tablist">
            <button id="tabCount" type="button" onclick="switchTab('count')" role="tab" aria-selected="true" class="flex-1 py-2 text-sm font-semibold rounded-md bg-white text-zinc-900 shadow-sm transition">Count</button>
            <button id="tabHistory" type="button" onclick="switchTab('history')" role="tab" aria-selected="false" class="flex-1 py-2 text-sm font-semibold rounded-md text-zinc-500 transition">Riwayat</button>
        </div>

        <div class="lg:grid lg:grid-cols-2 lg:gap-6 lg:items-start">

            <!-- Count panel -->
            <div id="panelCount" class="space-y-4 relative z-0">

                <!-- Step 1: Petugas -->
                <section id="step1Card" class="step-card step-card--active relative z-50">
                    <div class="step-card__head step-card__head--active">
                        <span class="step-card__badge step-card__badge--active">Step 1</span>
                        <h2 class="step-card__title step-card__title--active">Petugas</h2>
                    </div>
                    <div class="p-4">
                        <label for="counterName" class="block text-sm font-medium text-zinc-700 mb-1.5">Nama petugas</label>
                        <div class="flex gap-2">
                            <input type="text" id="counterName" name="counterName" autocomplete="off" inputmode="text" placeholder="Ketik nama atau scan ID badge" class="opname-field flex-1 min-w-0 border border-amber-200 p-3 rounded-lg bg-amber-50 text-amber-900 focus:border-violet-500 focus:ring-2 focus:ring-violet-500/20 focus:outline-none">
                            <button type="button" id="scanCounterBtn" class="shrink-0 px-4 py-3 rounded-lg bg-violet-600 hover:bg-violet-700 active:bg-violet-800 text-white text-sm font-semibold disabled:opacity-40">Scan</button>
                        </div>
                        <p class="text-xs text-zinc-500 mt-2">Scan QR pada ID badge, atau ketik nama.</p>
                    </div>
                </section>

                <!-- Step 2: Lokasi -->
                <section id="step2Card" class="step-card step-locked" aria-disabled="true">
                    <div id="step2CardHead" class="step-card__head step-card__head--idle">
                        <span class="step-card__badge step-card__badge--idle">Step 2</span>
                        <h2 class="step-card__title step-card__title--idle">Lokasi</h2>
                    </div>
                    <div class="p-4">
                        <label for="location" class="block text-sm font-medium text-zinc-700 mb-1.5">Kode lokasi</label>
                        <div class="flex gap-2">
                            <input type="text" id="location" readonly tabindex="-1" placeholder="Scan kode QR lokasi" aria-describedby="locationHint" class="opname-field flex-1 min-w-0 border border-amber-200 p-3 rounded-lg bg-amber-50 text-amber-900">
                            <button type="button" id="scanLocationBtn" disabled class="shrink-0 px-4 py-3 rounded-lg bg-violet-600 hover:bg-violet-700 active:bg-violet-800 text-white text-sm font-semibold disabled:opacity-40 disabled:cursor-not-allowed">Scan</button>
                        </div>
                        <p id="locationHint" class="text-xs text-zinc-500 mt-2">Pastikan lokasi terscan dahulu sebelum scan produk.</p>
                    </div>
                </section>

                <!-- Step 3: Produk -->
                <section id="step3Card" class="step-card step-locked" aria-disabled="true">
                    <div id="step3CardHead" class="step-card__head step-card__head--idle">
                        <span class="step-card__badge step-card__badge--idle">Step 3</span>
                        <h2 class="step-card__title step-card__title--idle">Produk</h2>
                    </div>
                    <div class="p-4 space-y-3">
                        <div class="sku-field-wrap">
                            <label for="skuInput" class="block text-sm font-medium text-zinc-700 mb-1.5">Kode SKU</label>
                            <div class="flex gap-2">
                                <input type="text" id="skuInput" autocomplete="off" disabled
                                    placeholder="Scan lokasi dulu"
                                    class="opname-field flex-1 min-w-0 border border-zinc-200 p-3 rounded-lg bg-zinc-50 text-zinc-400 focus:border-violet-500 focus:ring-2 focus:ring-violet-500/20 focus:outline-none">
                                <button type="button" id="scanSkuBtn" onclick="openScanModal('sku')" disabled
                                    class="shrink-0 px-4 py-3 rounded-lg bg-violet-600 hover:bg-violet-700 active:bg-violet-800 text-white text-sm font-semibold disabled:opacity-40 disabled:cursor-not-allowed">
                                    Scan
                                </button>
                            </div>
                            <ul id="skuSuggestions" class="hidden" role="listbox" aria-label="Saran SKU"></ul>
                            <p id="skuHint" class="text-xs text-zinc-500 mt-2">Ketik untuk saran otomatis atau scan QR SKU.</p>
                        </div>
                    </div>
                </section>

                <!-- Step 4: Jumlah -->
                <section id="step4Card" class="step-card step-locked" aria-disabled="true">
                    <div id="step4CardHead" class="step-card__head step-card__head--idle">
                        <span class="step-card__badge step-card__badge--idle">Step 4</span>
                        <h2 class="step-card__title step-card__title--idle">Jumlah</h2>
                    </div>
                    <div class="p-4 space-y-4">
                        <div class="flex items-center justify-center gap-2 sm:gap-3">
                            <button type="button" onclick="adjustCount(-10)" class="text-xs font-semibold text-zinc-600 bg-zinc-100 hover:bg-zinc-200 px-2.5 py-2 rounded-lg">−10</button>
                            <button type="button" onclick="adjustCount(-1)" class="flex h-11 w-11 sm:h-12 sm:w-12 items-center justify-center rounded-full bg-zinc-100 hover:bg-zinc-200 text-xl font-bold text-zinc-800">−</button>
                            <input type="number" id="count" oninput="updateSubmitState()" onfocus="this.select()" onclick="this.select()" class="w-24 sm:w-28 border border-zinc-200 rounded-xl text-center text-3xl sm:text-4xl font-bold tabular-nums text-zinc-900 focus:border-violet-500 focus:ring-2 focus:ring-violet-500/20 focus:outline-none" value="0" min="0">
                            <button type="button" onclick="adjustCount(1)" class="flex h-11 w-11 sm:h-12 sm:w-12 items-center justify-center rounded-full bg-violet-600 hover:bg-violet-700 text-xl font-bold text-white">+</button>
                            <button type="button" onclick="adjustCount(10)" class="text-xs font-semibold text-zinc-600 bg-zinc-100 hover:bg-zinc-200 px-2.5 py-2 rounded-lg">+10</button>
                        </div>
                        <div>
                            <label for="notes" class="block text-sm font-medium text-zinc-700 mb-1">Catatan <span class="text-zinc-400 font-normal">(opsional)</span></label>
                            <input type="text" id="notes" class="w-full border border-zinc-200 p-3 rounded-lg text-sm text-zinc-900 focus:border-violet-500 focus:ring-2 focus:ring-violet-500/20 focus:outline-none" placeholder="Contoh: kemasan rusak">
                        </div>
                    </div>
                </section>
            </div>

            <!-- History panel -->
            <div id="panelHistory" class="hidden lg:block">
                <div class="step-card opname-history-sticky lg:sticky">
                    <div class="step-card__head step-card__head--idle">
                        <h2 class="step-card__title step-card__title--idle flex-1">Riwayat</h2>
                        <button type="button" onclick="fetchHistory(true)" class="text-sm font-semibold text-violet-600 hover:text-violet-800">Refresh</button>
                    </div>
                    <div id="historyContainer" class="p-4 pt-2 space-y-2 max-h-[calc(100vh-12rem)] overflow-y-auto text-sm">
                        <p class="text-zinc-400 text-center py-8">Memuat riwayat…</p>
                    </div>
                </div>
            </div>
        </div>
    </main>

    <div id="footerSubmit" class="fixed bottom-0 left-0 right-0 z-20 bg-white/95 backdrop-blur-md border-t border-zinc-200 p-4 pb-[max(1rem,env(safe-area-inset-bottom))] lg:max-w-5xl lg:mx-auto lg:left-0 lg:right-0">
        <button id="submitBtn" type="button" onclick="submitData()" disabled class="w-full max-w-md mx-auto block bg-emerald-600 hover:bg-emerald-700 disabled:bg-zinc-300 disabled:cursor-not-allowed text-white font-semibold text-base py-3.5 rounded-xl shadow-sm active:scale-[0.98] transition flex items-center justify-center gap-2">
                        <span id="submitBtnLabel">Kirim ke sheet</span>
            <svg id="submitSpinner" class="hidden h-5 w-5 animate-spin" xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24"><circle class="opacity-25" cx="12" cy="12" r="10" stroke="currentColor" stroke-width="4"></circle><path class="opacity-75" fill="currentColor" d="M4 12a8 8 0 018-8V0C5.373 0 0 5.373 0 12h4z"></path></svg>
        </button>
    </div>

    <!-- Scanner modal (native hidden until opened — do not rely on Tailwind "hidden" alone) -->
    <div id="scanModal" hidden class="fixed inset-0 z-50" role="dialog" aria-modal="true" aria-labelledby="scanModalTitle">
        <div class="absolute inset-0 bg-black/70" id="scanModalBackdrop"></div>
        <div class="absolute inset-x-0 bottom-0 max-h-[92vh] flex flex-col bg-white rounded-t-2xl shadow-2xl lg:inset-auto lg:top-1/2 lg:left-1/2 lg:-translate-x-1/2 lg:-translate-y-1/2 lg:w-full lg:max-w-lg lg:rounded-2xl">
            <div class="flex items-center justify-between px-4 py-3 border-b border-zinc-100 shrink-0">
                <h3 id="scanModalTitle" class="text-base font-semibold text-zinc-900">Scan QR code</h3>
                <button type="button" onclick="closeScanModal()" class="text-sm font-medium text-zinc-500 hover:text-zinc-800 px-2 py-1">Cancel</button>
            </div>
            <div class="p-4 overflow-hidden flex-1 min-h-0">
                <div id="reader" class="w-full aspect-square max-h-[55vh] mx-auto rounded-xl overflow-hidden bg-black"></div>
                <p id="scanModalHint" class="text-center text-xs text-zinc-500 mt-3">Arahkan kamera ke QR code</p>
            </div>
        </div>
    </div>

    <script id="app-boot-data" type="application/json">{{ boot_data|tojson }}</script>
    <script src="https://unpkg.com/html5-qrcode@2.3.8/html5-qrcode.min.js"></script>
    <script>
        const SESSION_STORAGE_KEY = 'aeris_opname_session';

        function loadStoredSession() {
            try {
                const raw = localStorage.getItem(SESSION_STORAGE_KEY);
                return raw ? JSON.parse(raw) : null;
            } catch (e) { return null; }
        }

        function saveStoredSession(sessionId, sessionName) {
            localStorage.setItem(SESSION_STORAGE_KEY, JSON.stringify({
                sessionId,
                sessionName,
                startedAt: Date.now(),
            }));
        }

        function endSession() {
            localStorage.removeItem(SESSION_STORAGE_KEY);
            window.location.href = '/';
        }

        const activeSession = loadStoredSession();
        if (!activeSession || !activeSession.sessionId) {
            window.location.replace('/');
        }

        function sessionQueryParam() {
            return 'session_id=' + encodeURIComponent(activeSession.sessionId);
        }

        const _boot = (function () {
            try {
                return JSON.parse(document.getElementById('app-boot-data').textContent || '{}');
            } catch (e) {
                console.error('Boot data parse failed', e);
                return {};
            }
        })();
        let skuLookup = {};
        let skuCodes = [];
        let skuIndexLoaded = false;
        let skuIndexLoading = null;
        let skuSuggestTimer = null;
        const SKU_SUGGEST_MAX = 20;
        let locationLookup = _boot.location_lookup || {};
        let counterLookup = _boot.counter_lookup || {};
        let lookupWarnings = _boot.lookup_warnings || [];
        let locationAssignments = _boot.assignments || {};
        let enforcedSessionIds = new Set(_boot.enforced_session_ids || []);
        let validLocations = new Set(Object.values(locationLookup));
        let validCounters = new Set(Object.values(counterLookup));
        let currentTarget = '';
        let locationFrozen = false;
        let scannerRunning = false;
        let lastHandledScan = { key: '', at: 0, target: '' };
        let html5QrcodeScanner = null;
        const COUNTER_PREFIXES = [
            'counter:', 'counter name:', 'name:', 'nama:', 'nama petugas:',
            'petugas:', 'id:', 'badge:', 'id badge:',
        ];
        const LOCATION_PREFIXES = [
            'loc:', 'location:', 'lokasi:', 'kode lokasi:', 'precise location:',
        ];

        const FIELD_AMBER = "opname-field flex-1 min-w-0 border border-amber-200 p-3 rounded-lg bg-amber-50 text-amber-900 focus:border-violet-500 focus:ring-2 focus:ring-violet-500/20 focus:outline-none";
        const FIELD_EMERALD = "opname-field flex-1 min-w-0 border border-emerald-200 p-3 rounded-lg bg-emerald-50 text-emerald-900 focus:border-violet-500 focus:ring-2 focus:ring-violet-500/20 focus:outline-none";
        const CLS = {
            counterLocked: FIELD_AMBER,
            counterUnlocked: FIELD_EMERALD,
            locLocked: FIELD_AMBER,
            locUnlocked: FIELD_EMERALD,
            selLocked: "w-full border border-zinc-200 p-3 rounded-lg bg-zinc-50 text-zinc-400 text-sm font-medium",
            selUnlocked: "w-full border border-zinc-200 p-3 rounded-lg bg-white text-zinc-900 text-sm font-medium focus:border-violet-500 focus:ring-2 focus:ring-violet-500/20 focus:outline-none",
        };

        const STEP_DOT = {
            pending: "shrink-0 flex h-8 w-8 items-center justify-center rounded-full text-xs font-bold border-2 border-zinc-200 bg-white text-zinc-400",
            active: "shrink-0 flex h-8 w-8 items-center justify-center rounded-full text-xs font-bold border-2 border-violet-600 bg-violet-600 text-white",
            done: "shrink-0 flex h-8 w-8 items-center justify-center rounded-full text-xs font-bold border-2 border-emerald-500 bg-emerald-500 text-white",
        };

        function showLookupWarnings(warnings) {
            const box = document.getElementById('lookupWarnings');
            const inner = box.querySelector('div');
            if (!warnings || !warnings.length) {
                box.classList.add('hidden');
                inner.innerHTML = '';
                return;
            }
            inner.innerHTML = '<p class="font-semibold">Periksa sheet master:</p>' +
                warnings.map(w => `<p>• ${w}</p>`).join('');
            box.classList.remove('hidden');
        }

        function applyLookupData(data) {
            locationLookup = data.locations || {};
            counterLookup = data.counters || {};
            locationAssignments = data.assignments || {};
            enforcedSessionIds = new Set(data.enforced_session_ids || []);
            validLocations = new Set(Object.values(locationLookup));
            validCounters = new Set(Object.values(counterLookup));
            if (data.warnings) {
                lookupWarnings = data.warnings;
                showLookupWarnings(lookupWarnings);
            }
        }

        function isAssignmentEnforcedForSession() {
            return !!(activeSession && activeSession.sessionId && enforcedSessionIds.has(activeSession.sessionId));
        }

        function getAllowedLocationsForCounter(counterName) {
            if (!isAssignmentEnforcedForSession()) return null;
            const resolved = resolveCounter(counterName);
            if (!resolved) return [];
            const sess = locationAssignments[activeSession.sessionId] || {};
            return sess[resolved.toLowerCase()] || [];
        }

        function assignmentBlockMessage(counterName) {
            const allowed = getAllowedLocationsForCounter(counterName);
            const resolved = resolveCounter(counterName) || counterName;
            if (!allowed || !allowed.length) {
                return `Petugas ${resolved} belum ditugaskan lokasi untuk sesi ini. Hubungi admin.`;
            }
            let preview = allowed.slice(0, 8).join(', ');
            if (allowed.length > 8) preview += `, … (+${allowed.length - 8} lokasi)`;
            return `Lokasi tidak ditugaskan untuk ${resolved}. Lokasi Anda: ${preview}.`;
        }

        function isLocationAllowedForPetugas(counterName, location) {
            const allowed = getAllowedLocationsForCounter(counterName);
            if (allowed === null) return true;
            if (!allowed.length) return false;
            return allowed.includes(location);
        }

        function resolveSku(code) {
            const trimmed = String(code || '').trim();
            if (!trimmed || !Object.keys(skuLookup).length) return null;
            if (Object.values(skuLookup).includes(trimmed)) return trimmed;
            const key = normalizeScanText(trimmed).toLowerCase();
            if (key && skuLookup[key]) return skuLookup[key];
            const plain = trimmed.toLowerCase().replace(/\\s+/g, ' ');
            if (plain && skuLookup[plain]) return skuLookup[plain];
            return null;
        }

        async function ensureSkuIndex() {
            if (skuIndexLoaded) return true;
            if (skuIndexLoading) return skuIndexLoading;
            skuIndexLoading = (async () => {
                try {
                    const res = await fetch('/api/sku-codes');
                    if (!res.ok) throw new Error('fetch failed');
                    const data = await res.json();
                    skuLookup = data.sku_lookup || {};
                    skuCodes = data.sku_codes || [];
                    if (data.warnings && data.warnings.length) {
                        lookupWarnings = lookupWarnings.concat(data.warnings);
                        showLookupWarnings(lookupWarnings);
                    }
                    skuIndexLoaded = true;
                    return true;
                } catch (e) {
                    showToast('Daftar SKU gagal dimuat. Periksa tab SKU List.', 'error');
                    return false;
                } finally {
                    skuIndexLoading = null;
                }
            })();
            return skuIndexLoading;
        }

        function hideSkuSuggestions() {
            const box = document.getElementById('skuSuggestions');
            box.classList.add('hidden');
            box.innerHTML = '';
            const step3 = document.getElementById('step3Card');
            if (step3) step3.classList.remove('sku-suggest-open');
        }

        function filterSkuSuggestions(query) {
            const q = String(query || '').trim().toLowerCase();
            if (!q || !skuCodes.length) return [];
            const matches = [];
            for (const code of skuCodes) {
                if (code.toLowerCase().includes(q)) {
                    matches.push(code);
                    if (matches.length >= SKU_SUGGEST_MAX) break;
                }
            }
            return matches;
        }

        function renderSkuSuggestions(matches) {
            const box = document.getElementById('skuSuggestions');
            if (!matches.length) {
                hideSkuSuggestions();
                return;
            }
            box.innerHTML = matches.map((sku, i) =>
                `<li role="option" tabindex="-1"${i === 0 ? ' aria-selected="true"' : ''}>${escapeHtml(sku)}</li>`
            ).join('');
            box.classList.remove('hidden');
            const step3 = document.getElementById('step3Card');
            if (step3) step3.classList.add('sku-suggest-open');
            box.querySelectorAll('li').forEach((li, i) => {
                li.addEventListener('mousedown', (e) => {
                    e.preventDefault();
                    selectSkuSuggestion(matches[i]);
                });
            });
        }

        function selectSkuSuggestion(sku) {
            const input = document.getElementById('skuInput');
            input.value = sku;
            hideSkuSuggestions();
            setSkuInputStyle('valid');
            updateStepperUI();
            updateSubmitState();
        }

        function setSkuInputStyle(state) {
            const input = document.getElementById('skuInput');
            if (input.disabled) {
                input.className = 'opname-field flex-1 min-w-0 border border-zinc-200 p-3 rounded-lg bg-zinc-50 text-zinc-400 focus:border-violet-500 focus:ring-2 focus:ring-violet-500/20 focus:outline-none';
                return;
            }
            if (state === 'valid') {
                input.className = FIELD_EMERALD;
            } else if (state === 'invalid') {
                input.className = 'opname-field flex-1 min-w-0 border border-rose-300 p-3 rounded-lg bg-rose-50 text-rose-900 focus:border-violet-500 focus:ring-2 focus:ring-violet-500/20 focus:outline-none';
            } else {
                input.className = FIELD_AMBER;
            }
        }

        function onSkuInput() {
            const input = document.getElementById('skuInput');
            setSkuInputStyle(false);
            clearTimeout(skuSuggestTimer);
            skuSuggestTimer = setTimeout(() => {
                renderSkuSuggestions(filterSkuSuggestions(input.value));
            }, 150);
            updateStepperUI();
            updateSubmitState();
        }

        function commitSkuInput() {
            const input = document.getElementById('skuInput');
            const trimmed = input.value.trim();
            if (!trimmed) {
                setSkuInputStyle(false);
                hideSkuSuggestions();
                updateStepperUI();
                updateSubmitState();
                return;
            }
            const resolved = resolveSku(trimmed);
            if (!resolved) {
                setSkuInputStyle('invalid');
                updateStepperUI();
                updateSubmitState();
                return;
            }
            input.value = resolved;
            setSkuInputStyle('valid');
            hideSkuSuggestions();
            updateStepperUI();
            updateSubmitState();
        }

        function onSkuBlur() {
            setTimeout(hideSkuSuggestions, 200);
            commitSkuInput();
        }

        function applySkuScan(text) {
            const resolved = resolveSku(text) || resolveSku(normalizeScanText(text));
            if (!resolved) {
                const scanned = normalizeScanText(text) || String(text).trim();
                showToast(`SKU tidak dikenali: "${scanned.slice(0, 40)}".`, 'warning');
                setSkuInputStyle('invalid');
                return false;
            }
            document.getElementById('skuInput').value = resolved;
            setSkuInputStyle('valid');
            hideSkuSuggestions();
            updateStepperUI();
            updateSubmitState();
            return true;
        }

        async function loadLookups() {
            try {
                const res = await fetch('/api/lookups');
                if (res.ok) {
                    applyLookupData(await res.json());
                }
            } catch (e) {}
        }

        function resetScanDebounce() {
            lastHandledScan = { key: '', at: 0, target: '' };
        }

        function normalizeScanText(code, kind = 'any') {
            let s = String(code || '').trim().replace(/\\ufeff/g, '').replace(/\\u200b/g, '').replace(/\\r/g, '').replace(/\\n/g, '');
            if (!s) return '';
            if (/^https?:\\/\\//i.test(s) || s.includes('://')) {
                s = s.replace(/\\/+$/, '').split('/').pop();
            }
            if (s.includes('?')) s = s.split('?')[0];
            if (s.includes('#')) s = s.split('#')[0];
            s = s.trim();
            let prefixes = COUNTER_PREFIXES.concat(LOCATION_PREFIXES);
            if (kind === 'counter') prefixes = COUNTER_PREFIXES;
            else if (kind === 'location') prefixes = LOCATION_PREFIXES;
            for (const p of prefixes) {
                if (s.toLowerCase().startsWith(p)) {
                    s = s.slice(p.length).trim();
                    break;
                }
            }
            return s.trim();
        }

        function scanKindForTarget(target) {
            if (target === 'counter') return 'counter';
            if (target === 'location') return 'location';
            return 'any';
        }

        function shouldHandleScan(raw, target) {
            const kind = scanKindForTarget(target || currentTarget);
            const key = normalizeScanText(raw, kind).toLowerCase();
            if (!key) return false;
            const now = Date.now();
            if (lastHandledScan.key === key && lastHandledScan.target === target && now - lastHandledScan.at < 500) {
                return false;
            }
            return true;
        }

        function markScanHandled(raw, target) {
            const kind = scanKindForTarget(target || currentTarget);
            const key = normalizeScanText(raw, kind).toLowerCase();
            if (key) lastHandledScan = { key, at: Date.now(), target: target || currentTarget };
        }

        function syncTopChromeLayout() {
            const body = document.body;
            const chrome = document.getElementById('opnameTopChrome');
            const bar = document.getElementById('locationStickyBar');
            const locked = locationFrozen && bar && !bar.classList.contains('hidden');
            if (locked && chrome) {
                body.classList.add('opname-location-locked');
                const h = Math.ceil(chrome.getBoundingClientRect().height);
                body.style.setProperty('--opname-chrome-h', h + 'px');
            } else {
                body.classList.remove('opname-location-locked');
                body.style.removeProperty('--opname-chrome-h');
            }
        }

        function updateLocationStickyBar() {
            const bar = document.getElementById('locationStickyBar');
            const valEl = document.getElementById('locationStickyValue');
            if (!bar || !valEl) return;
            const loc = resolveLocation(document.getElementById('location').value);
            if (locationFrozen && loc) {
                valEl.textContent = loc;
                bar.classList.remove('hidden');
            } else {
                bar.classList.add('hidden');
            }
            requestAnimationFrame(syncTopChromeLayout);
        }

        function freezeLocation() {
            locationFrozen = true;
            document.getElementById('scanLocationBtn').disabled = true;
            updateLocationUI();
            updateLocationStickyBar();
        }

        function unfreezeLocation() {
            locationFrozen = false;
            const counterOk = isValidCounter(document.getElementById('counterName').value);
            document.getElementById('scanLocationBtn').disabled = !counterOk;
            updateLocationUI();
            updateLocationStickyBar();
        }

        function requestChangeLocation() {
            if (!locationFrozen) return;
            if (!confirm('Ganti lokasi? Anda perlu scan QR lokasi baru sebelum menghitung produk.')) return;
            const locInput = document.getElementById('location');
            locInput.value = '';
            locInput.className = CLS.locLocked;
            unfreezeLocation();
            lockSkuFields();
            updateStepperUI();
            updateSubmitState();
        }

        function updateLocationUI() {
            const locInput = document.getElementById('location');
            const hint = document.getElementById('locationHint');
            const counterOk = isValidCounter(document.getElementById('counterName').value);
            if (!counterOk) {
                if (hint) hint.textContent = 'Pastikan lokasi terscan dahulu sebelum scan produk.';
                locInput.placeholder = 'Scan kode QR lokasi';
            } else if (locInput.value.trim()) {
                if (locationFrozen) {
                    if (hint) hint.textContent = 'Pastikan lokasi sudah benar.';
                } else if (hint) {
                    hint.textContent = 'Lokasi terisi. Tekan Scan untuk mengganti.';
                }
                locInput.placeholder = locInput.value;
            } else {
                if (hint) hint.textContent = 'Scan kode QR lokasi sebelum scan produk.';
                locInput.placeholder = 'Scan kode QR lokasi';
            }
            updateLocationStickyBar();
        }

        function lockSkuFields() {
            document.getElementById('scanSkuBtn').disabled = true;
            const input = document.getElementById('skuInput');
            input.disabled = true;
            input.value = '';
            input.placeholder = 'Scan lokasi dulu';
            setSkuInputStyle(false);
            hideSkuSuggestions();
        }

        function resetLocationAndSku() {
            const locInput = document.getElementById('location');
            locInput.value = '';
            locInput.className = CLS.locLocked;
            unfreezeLocation();
            lockSkuFields();
            updateStepperUI();
            updateSubmitState();
        }

        async function enableSkuFields() {
            document.getElementById('scanSkuBtn').disabled = false;
            const input = document.getElementById('skuInput');
            input.disabled = false;
            input.value = '';
            input.placeholder = 'Memuat daftar SKU…';
            setSkuInputStyle(false);
            const ok = await ensureSkuIndex();
            input.placeholder = ok ? 'Ketik atau scan kode SKU' : 'SKU tidak tersedia';
            if (!ok) {
                document.getElementById('scanSkuBtn').disabled = true;
                input.disabled = true;
            }
        }

        function resetSkuAndCount() {
            document.getElementById('count').value = '0';
            document.getElementById('notes').value = '';
            const input = document.getElementById('skuInput');
            input.value = '';
            hideSkuSuggestions();
            if (!input.disabled) {
                setSkuInputStyle(false);
                input.placeholder = 'Ketik atau scan kode SKU';
            }
            updateStepperUI();
            updateSubmitState();
        }

        async function syncUIState(opts) {
            const refreshHistory = opts && opts.refreshHistory;
            const counterInput = document.getElementById('counterName');
            const locInput = document.getElementById('location');
            const historyContainer = document.getElementById('historyContainer');

            if (isValidCounter(counterInput.value)) {
                counterInput.className = CLS.counterUnlocked;
                counterInput.removeAttribute('readonly');
                if (!locationFrozen) {
                    document.getElementById('scanLocationBtn').disabled = false;
                }
                setStepCardEnabled('step2Card', true);
                updateLocationUI();
                if (locInput.value.trim() && isValidLocation(locInput.value)) {
                    locInput.className = CLS.locUnlocked;
                    if (!locationFrozen) {
                        unlockFormForLocation();
                    } else {
                        if (document.getElementById('skuInput').disabled) {
                            enableSkuFields();
                        } else {
                            document.getElementById('scanSkuBtn').disabled = false;
                        }
                        updateLocationUI();
                    }
                } else {
                    if (locInput.value.trim() && !isValidLocation(locInput.value)) {
                        locInput.value = '';
                    }
                    if (locationFrozen && !locInput.value.trim()) {
                        locationFrozen = false;
                    }
                    unfreezeLocation();
                    lockSkuFields();
                    locInput.className = CLS.locLocked;
                }
                if (refreshHistory) maybeFetchHistory();
            } else {
                counterInput.className = CLS.counterLocked;
                counterInput.removeAttribute('readonly');
                document.getElementById('scanLocationBtn').disabled = true;
                setStepCardEnabled('step2Card', false);
                locInput.value = '';
                locInput.className = CLS.locLocked;
                unfreezeLocation();
                lockSkuFields();
                historyContainer.innerHTML = '<p class="text-zinc-400 text-center py-8">Scan badge atau ketik nama petugas.</p>';
            }
            updateStepperUI();
            updateSubmitState();
        }

        function getScanner() {
            if (!html5QrcodeScanner && typeof Html5Qrcode !== 'undefined') {
                html5QrcodeScanner = new Html5Qrcode('reader');
            }
            return html5QrcodeScanner;
        }

        function bindScanButtons() {
            const counterBtn = document.getElementById('scanCounterBtn');
            const locationBtn = document.getElementById('scanLocationBtn');
            const counterInput = document.getElementById('counterName');
            if (counterBtn) {
                counterBtn.addEventListener('click', (e) => {
                    e.preventDefault();
                    e.stopPropagation();
                    openScanModal('counter');
                });
            }
            if (locationBtn) {
                locationBtn.addEventListener('click', (e) => {
                    e.preventDefault();
                    e.stopPropagation();
                    openScanModal('location');
                });
            }
            if (counterInput) {
                counterInput.addEventListener('input', onCounterNameInput);
                counterInput.addEventListener('blur', onCounterNameInput);
                counterInput.addEventListener('keydown', (e) => {
                    if (e.key === 'Enter') onCounterNameInput(e);
                });
            }
            const changeLocBtn = document.getElementById('changeLocationBtn');
            if (changeLocBtn) {
                changeLocBtn.addEventListener('click', (e) => {
                    e.preventDefault();
                    requestChangeLocation();
                });
            }
            const backdrop = document.getElementById('scanModalBackdrop');
            if (backdrop) {
                backdrop.addEventListener('click', () => closeScanModal());
            }
        }

        function resetPageInteractionState() {
            document.body.style.overflow = '';
            document.body.style.pointerEvents = '';
            const modal = document.getElementById('scanModal');
            if (!modal) return;
            modal.hidden = true;
            modal.classList.remove('is-open');
            modal.style.display = 'none';
        }

        async function initApp() {
            const badge = document.getElementById('sessionBadge');
            if (badge && activeSession) {
                badge.textContent = activeSession.sessionName || activeSession.sessionId;
                badge.title = activeSession.sessionName || activeSession.sessionId;
            }
            resetPageInteractionState();
            bindScanButtons();
            const skuInput = document.getElementById('skuInput');
            if (skuInput) {
                skuInput.addEventListener('input', onSkuInput);
                skuInput.addEventListener('blur', onSkuBlur);
                skuInput.addEventListener('keydown', (e) => {
                    if (e.key === 'Escape') hideSkuSuggestions();
                    if (e.key === 'Enter') {
                        e.preventDefault();
                        commitSkuInput();
                        hideSkuSuggestions();
                    }
                });
            }
            showLookupWarnings(lookupWarnings);
            await syncUIState({ refreshHistory: false });
            syncTopChromeLayout();
            window.addEventListener('resize', () => {
                if (locationFrozen) syncTopChromeLayout();
            });
            switchTab('count');
        }

        if (document.readyState === 'loading') {
            document.addEventListener('DOMContentLoaded', () => { initApp().catch(console.error); });
        } else {
            initApp().catch(console.error);
        }

        function switchTab(tab) {
            const isLg = window.matchMedia('(min-width: 1024px)').matches;
            const panelCount = document.getElementById('panelCount');
            const panelHistory = document.getElementById('panelHistory');
            const footer = document.getElementById('footerSubmit');
            const tabCount = document.getElementById('tabCount');
            const tabHistory = document.getElementById('tabHistory');

            if (isLg) {
                panelCount.classList.remove('hidden');
                panelHistory.classList.remove('hidden');
                footer.classList.remove('hidden');
                return;
            }

            const showCount = tab === 'count';
            panelCount.classList.toggle('hidden', !showCount);
            panelHistory.classList.toggle('hidden', showCount);
            footer.classList.toggle('hidden', !showCount);

            tabCount.className = showCount
                ? 'flex-1 py-2 text-sm font-semibold rounded-md bg-white text-zinc-900 shadow-sm transition'
                : 'flex-1 py-2 text-sm font-semibold rounded-md text-zinc-500 transition';
            tabHistory.className = showCount
                ? 'flex-1 py-2 text-sm font-semibold rounded-md text-zinc-500 transition'
                : 'flex-1 py-2 text-sm font-semibold rounded-md bg-white text-zinc-900 shadow-sm transition';
            tabCount.setAttribute('aria-selected', showCount);
            tabHistory.setAttribute('aria-selected', !showCount);
            if (!showCount) maybeFetchHistory();
        }

        function isHistoryPanelVisible() {
            return window.matchMedia('(min-width: 1024px)').matches
                || document.getElementById('tabHistory').getAttribute('aria-selected') === 'true';
        }

        function maybeFetchHistory(forceRefresh = false) {
            if (!isHistoryPanelVisible()) return;
            if (!isValidCounter(document.getElementById('counterName').value)) return;
            fetchHistory(forceRefresh);
        }

        window.addEventListener('resize', () => {
            switchTab(document.getElementById('tabHistory').getAttribute('aria-selected') === 'true' ? 'history' : 'count');
            maybeFetchHistory();
        });

        function setStepCardVisual(cardId, active) {
            const card = document.getElementById(cardId);
            const head = document.getElementById(cardId + 'Head');
            if (!card) return;
            card.classList.toggle('step-card--active', active);
            if (head) {
                head.classList.toggle('step-card__head--active', active);
                head.classList.toggle('step-card__head--idle', !active);
                const badge = head.querySelector('.step-card__badge');
                const title = head.querySelector('.step-card__title');
                if (badge) {
                    badge.classList.toggle('step-card__badge--active', active);
                    badge.classList.toggle('step-card__badge--idle', !active);
                }
                if (title) {
                    title.classList.toggle('step-card__title--active', active);
                    title.classList.toggle('step-card__title--idle', !active);
                }
            }
        }

        function setStepCardEnabled(cardId, enabled) {
            if (cardId === 'step1Card') return;
            const card = document.getElementById(cardId);
            if (!card) return;
            card.classList.toggle('step-locked', !enabled);
            card.setAttribute('aria-disabled', String(!enabled));
            setStepCardVisual(cardId, enabled);
        }

        function setStepDot(num, state) {
            document.getElementById(`step${num}Dot`).className = STEP_DOT[state];
            const label = document.getElementById(`step${num}Indicator`).querySelector('span:last-child');
            if (label) {
                label.className = state === 'pending'
                    ? 'text-xs font-medium text-zinc-400 truncate hidden sm:inline'
                    : 'text-xs font-medium text-zinc-700 truncate hidden sm:inline';
            }
        }

        function updateStepperUI() {
            const counterOk = isValidCounter(document.getElementById('counterName').value);
            const loc = document.getElementById('location').value.trim();
            const sku = document.getElementById('skuInput').value.trim();

            setStepCardEnabled('step2Card', counterOk);

            if (!counterOk) {
                setStepDot(1, 'active');
                setStepDot(2, 'pending');
                setStepDot(3, 'pending');
                setStepDot(4, 'pending');
                setStepCardEnabled('step3Card', false);
                setStepCardEnabled('step4Card', false);
                return;
            }

            if (!loc) {
                setStepDot(1, 'done');
                setStepDot(2, 'active');
                setStepDot(3, 'pending');
                setStepDot(4, 'pending');
                setStepCardEnabled('step3Card', false);
                setStepCardEnabled('step4Card', false);
            } else if (!sku || !resolveSku(sku)) {
                setStepDot(1, 'done');
                setStepDot(2, 'done');
                setStepDot(3, 'active');
                setStepDot(4, 'pending');
                setStepCardEnabled('step3Card', true);
                setStepCardEnabled('step4Card', false);
            } else {
                setStepDot(1, 'done');
                setStepDot(2, 'done');
                setStepDot(3, 'done');
                setStepDot(4, 'active');
                setStepCardEnabled('step3Card', true);
                setStepCardEnabled('step4Card', true);
            }
        }

        function updateSubmitState() {
            const countVal = document.getElementById('count').value;
            const countOk = countVal !== '' && Number(countVal) > 0;
            const ready = isValidCounter(document.getElementById('counterName').value)
                && resolveLocation(document.getElementById('location').value)
                && resolveSku(document.getElementById('skuInput').value)
                && countOk;
            const btn = document.getElementById('submitBtn');
            btn.disabled = !ready || btn.dataset.loading === '1';
        }

        function resolveCounter(name) {
            const trimmed = String(name || '').trim();
            if (!trimmed) return null;
            const key = normalizeScanText(trimmed, 'counter').toLowerCase();
            if (key && counterLookup[key]) return counterLookup[key];
            const plain = trimmed.toLowerCase().replace(/\\s+/g, ' ');
            if (plain && counterLookup[plain]) return counterLookup[plain];
            if (validCounters.has(trimmed)) return trimmed;
            for (const c of validCounters) {
                if (String(c).toLowerCase() === plain) return c;
            }
            return null;
        }

        function isValidCounter(name) {
            return !!resolveCounter(name);
        }

        async function applyCounterScan(text) {
            if (!Object.keys(counterLookup).length) {
                await loadLookups();
            }
            if (!Object.keys(counterLookup).length) {
                showToast('Daftar petugas belum dimuat. Hubungi admin.', 'error');
                return false;
            }
            const resolved = resolveCounter(text);
            if (!resolved) {
                const scanned = normalizeScanText(text, 'counter') || String(text).trim();
                showToast(`Petugas tidak dikenali: "${scanned.slice(0, 40)}". Ketik nama sesuai dengan ID badge.`, 'warning');
                return false;
            }
            document.getElementById('counterName').value = resolved;
            unlockAfterCounter();
            await syncUIState({ refreshHistory: false });
            maybeFetchHistory();
            return true;
        }

        let lastCounterNameToast = '';
        async function onCounterNameInput(ev) {
            const counterInput = document.getElementById('counterName');
            const trimmed = String(counterInput.value || '').trim();
            const isCommit = ev && (ev.type === 'blur' || (ev.type === 'keydown' && ev.key === 'Enter'));

            if (!trimmed) {
                lastCounterNameToast = '';
                await syncUIState({ refreshHistory: false });
                return;
            }

            if (!Object.keys(counterLookup).length) {
                await loadLookups();
            }

            const resolved = resolveCounter(counterInput.value);
            if (!resolved) {
                if (isCommit) {
                    let msg;
                    let type = 'warning';
                    if (!Object.keys(counterLookup).length) {
                        msg = 'Daftar petugas belum dimuat. Hubungi admin.';
                        type = 'error';
                    } else {
                        const key = normalizeScanText(trimmed, 'counter') || trimmed;
                        msg = `Petugas tidak dikenali: "${String(key).slice(0, 40)}". Ketik nama sesuai dengan ID badge.`;
                    }
                    if (lastCounterNameToast !== msg) {
                        showToast(msg, type);
                        lastCounterNameToast = msg;
                    }
                }
            } else {
                lastCounterNameToast = '';
                if (isCommit && counterInput.value !== resolved) {
                    counterInput.value = resolved;
                }
            }
            await syncUIState({ refreshHistory: false });
            if (isCommit && resolved) maybeFetchHistory();
        }

        function unlockAfterCounter() {
            document.getElementById('counterName').className = CLS.counterUnlocked;
            if (!locationFrozen) {
                document.getElementById('scanLocationBtn').disabled = false;
            }
            updateLocationUI();
            updateStepperUI();
            updateSubmitState();
        }

        let toastTimer;
        function showToast(message, type = 'success') {
            const toast = document.getElementById('toast');
            const inner = document.getElementById('toastInner');
            const styles = {
                success: 'bg-emerald-50 text-emerald-900 border-emerald-200',
                error: 'bg-rose-50 text-rose-900 border-rose-200',
                warning: 'bg-amber-50 text-amber-900 border-amber-200',
            };
            inner.className = `rounded-xl px-4 py-3 text-sm font-medium shadow-lg border ${styles[type] || styles.success}`;
            inner.textContent = message;
            toast.classList.remove('translate-y-[-120%]', 'opacity-0');
            clearTimeout(toastTimer);
            toastTimer = setTimeout(() => {
                toast.classList.add('translate-y-[-120%]', 'opacity-0');
            }, type === 'warning' ? 6000 : 3500);
        }

        function resolveLocation(code) {
            const trimmed = String(code || '').trim();
            if (!trimmed) return null;
            const key = normalizeScanText(trimmed, 'location').toLowerCase();
            if (key && locationLookup[key]) return locationLookup[key];
            const plain = trimmed.toLowerCase().replace(/\\s+/g, ' ');
            if (plain && locationLookup[plain]) return locationLookup[plain];
            if (validLocations.has(trimmed)) return trimmed;
            for (const v of validLocations) {
                if (String(v).toLowerCase() === plain) return v;
            }
            return null;
        }

        function isValidLocation(code) {
            return !!resolveLocation(code);
        }

        function applyLocationScan(text) {
            if (!isValidCounter(document.getElementById('counterName').value)) {
                showToast('Scan kode QR lokasi', 'warning');
                return false;
            }
            if (!Object.keys(locationLookup).length) {
                showToast('Daftar lokasi belum dimuat. Hubungi admin.', 'error');
                return false;
            }
            const resolved = resolveLocation(text);
            if (!resolved) {
                const scanned = normalizeScanText(text, 'location') || String(text).trim();
                showToast(`Lokasi tidak dikenali: "${scanned.slice(0, 40)}". Periksa tab LOCATIONS.`, 'warning');
                return false;
            }
            const counterName = document.getElementById('counterName').value;
            if (!isLocationAllowedForPetugas(counterName, resolved)) {
                showToast(assignmentBlockMessage(counterName), 'warning');
                return false;
            }
            document.getElementById('location').value = resolved;
            unlockFormForLocation();
            return true;
        }

        function unlockFormForLocation() {
            const locInput = document.getElementById('location');
            locInput.className = CLS.locUnlocked;
            freezeLocation();
            enableSkuFields();
            updateStepperUI();
            updateSubmitState();
        }

        async function ensureScannerStopped() {
            const scanner = getScanner();
            if (!scanner) return;
            try {
                await scanner.stop();
            } catch (e) {}
            try {
                scanner.clear();
            } catch (e) {}
            scannerRunning = false;
        }

        async function openScanModal(target) {
            if (target === 'location' && locationFrozen) {
                showToast('Pastikan lokasi sudah benar.', 'warning');
                return;
            }
            if (target === 'location' && document.getElementById('scanLocationBtn').disabled) return;
            if (target === 'sku' && document.getElementById('scanSkuBtn').disabled) return;
            if (typeof Html5Qrcode === 'undefined') {
                showToast('Pemindai QR tidak termuat. Muat ulang halaman.', 'error');
                return;
            }

            currentTarget = target;
            const scanTitles = {
                counter: 'Scan ID badge',
                location: 'Scan lokasi',
                sku: 'Scan SKU',
            };
            const scanHints = {
                counter: 'Arahkan kamera ke QR pada ID badge petugas',
                location: 'Arahkan kamera ke QR lokasi',
                sku: 'Arahkan kamera ke QR SKU',
            };
            document.getElementById('scanModalTitle').textContent = scanTitles[target] || 'Scan QR';
            document.getElementById('scanModalHint').textContent = scanHints[target] || 'Arahkan kamera ke QR code';
            const modal = document.getElementById('scanModal');
            modal.hidden = false;
            modal.classList.add('is-open');
            modal.style.display = 'block';
            document.body.style.overflow = 'hidden';
            resetScanDebounce();

            await ensureScannerStopped();
            await new Promise(r => setTimeout(r, 200));

            const scanner = getScanner();
            if (!scanner) {
                showToast('Pemindai QR tidak siap. Muat ulang halaman.', 'error');
                closeScanModal();
                return;
            }

            try {
                scannerRunning = true;
                await scanner.start(
                    { facingMode: "environment" },
                    { fps: 15, qrbox: { width: 250, height: 250 } },
                    async (decodedText) => {
                        const text = decodedText.trim();
                        if (!text || !shouldHandleScan(text, currentTarget)) return;

                        let shouldClose = false;
                        if (currentTarget === 'counter') {
                            shouldClose = await applyCounterScan(text);
                        } else if (currentTarget === 'location') {
                            shouldClose = applyLocationScan(text);
                        } else if (currentTarget === 'sku') {
                            if (!skuIndexLoaded) await ensureSkuIndex();
                            shouldClose = applySkuScan(text);
                        }
                        if (shouldClose) {
                            markScanHandled(text, currentTarget);
                            await closeScanModal();
                        }
                    },
                    () => {}
                );
            } catch (err) {
                scannerRunning = false;
                showToast('Could not start camera. Check permissions.', 'error');
                closeScanModal();
            }
        }

        async function closeScanModal() {
            await ensureScannerStopped();
            const modal = document.getElementById('scanModal');
            modal.hidden = true;
            modal.classList.remove('is-open');
            modal.style.display = 'none';
            document.body.style.overflow = '';
        }

        window.openScanModal = openScanModal;
        window.closeScanModal = closeScanModal;
        window.onCounterNameInput = onCounterNameInput;
        window.switchTab = switchTab;
        window.submitData = submitData;
        window.fetchHistory = fetchHistory;
        window.adjustCount = adjustCount;

        function adjustCount(amount) {
            const countInput = document.getElementById('count');
            let currentVal = parseInt(countInput.value) || 0;
            currentVal += amount;
            if (currentVal < 0) currentVal = 0;
            countInput.value = currentVal;
            updateSubmitState();
        }

        function escapeHtml(str) {
            return String(str)
                .replace(/&/g, '&amp;')
                .replace(/</g, '&lt;')
                .replace(/>/g, '&gt;')
                .replace(/"/g, '&quot;');
        }

        async function fetchHistory(forceRefresh = false) {
            const counterName = resolveCounter(document.getElementById('counterName').value) || '';
            const container = document.getElementById('historyContainer');

            if (!counterName) {
                container.innerHTML = '<p class="text-zinc-400 text-center py-8">Scan badge atau ketik nama petugas.</p>';
                return;
            }

            container.innerHTML = '<p class="text-zinc-400 text-center py-8 animate-pulse">Memuat…</p>';

            try {
                const qs = new URLSearchParams({ name: counterName, session_id: activeSession.sessionId });
                if (forceRefresh) qs.set('refresh', '1');
                const response = await fetch(`/history?${qs.toString()}`);
                const payload = await response.json();
                const data = Array.isArray(payload) ? payload : (payload.items || []);
                const truncated = !Array.isArray(payload) && !!payload.truncated;

                if (data.length === 0) {
                    container.innerHTML = '<p class="text-zinc-400 text-center py-8">Belum ada catatan untuk petugas ini.</p>';
                    return;
                }

                const rowsHtml = data.map(item => `
                    <div class="border border-zinc-100 rounded-lg p-3 hover:bg-zinc-50/80 transition">
                        <div class="flex justify-between items-start gap-2">
                            <div class="min-w-0 flex-1">
                                <div class="flex justify-between items-baseline gap-2">
                                    <span class="text-sm font-semibold text-violet-700 truncate">${escapeHtml(item.location)}</span>
                                    <span class="text-xl font-bold tabular-nums text-zinc-900 shrink-0">${escapeHtml(String(item.count))}</span>
                                </div>
                                <p class="text-sm font-medium text-zinc-700 truncate mt-0.5">${escapeHtml(item.sku)}</p>
                                ${item.timestamp ? `<p class="text-xs text-zinc-400 mt-0.5">${escapeHtml(item.timestamp)}</p>` : ''}
                                ${item.notes ? `<p class="text-xs text-zinc-500 mt-0.5 truncate">${escapeHtml(item.notes)}</p>` : ''}
                            </div>
                        </div>
                        <div class="flex gap-2 mt-2 pt-2 border-t border-zinc-100">
                            <button type="button" onclick="editItem('${escapeHtml(item.id)}', ${parseInt(item.count) || 0})" class="text-xs font-medium text-amber-700 hover:text-amber-900">Edit</button>
                            <button type="button" onclick="deleteItem('${escapeHtml(item.id)}')" class="text-xs font-medium text-rose-600 hover:text-rose-800">Delete</button>
                        </div>
                    </div>
                `).join('');
                const truncNote = truncated
                    ? '<p class="text-xs text-amber-700 text-center py-2 px-2">Menampilkan entri terbaru saja. Ketuk Refresh jika perlu memuat ulang.</p>'
                    : '';
                container.innerHTML = rowsHtml + truncNote;
            } catch (err) {
                container.innerHTML = '<p class="text-rose-600 text-center py-8">Failed to load history.</p>';
            }
        }

        async function submitData() {
            const locInput = document.getElementById('location').value;
            const skuInput = document.getElementById('skuInput').value;
            const countInput = document.getElementById('count').value;
            const btn = document.getElementById('submitBtn');
            const label = document.getElementById('submitBtnLabel');
            const spinner = document.getElementById('submitSpinner');

            const counterName = resolveCounter(document.getElementById('counterName').value);
            const resolvedLocation = resolveLocation(locInput);
            const resolvedSku = resolveSku(skuInput);

            if (!counterName) {
                showToast('Scan ID badge petugas dulu.', 'warning');
                return;
            }

            if (!resolvedLocation || !resolvedSku || countInput === '') {
                showToast('Lengkapi lokasi, SKU, dan jumlah sebelum submit.', 'warning');
                return;
            }

            if (!isLocationAllowedForPetugas(document.getElementById('counterName').value, resolvedLocation)) {
                showToast(assignmentBlockMessage(document.getElementById('counterName').value), 'warning');
                return;
            }

            btn.disabled = true;
            btn.dataset.loading = '1';
            label.textContent = 'Menyimpan…';
            spinner.classList.remove('hidden');

            const payload = {
                session_id: activeSession.sessionId,
                counter_name: counterName,
                location: resolvedLocation,
                sku: resolvedSku,
                count: countInput,
                notes: document.getElementById('notes').value.trim()
            };

            try {
                const response = await fetch('/submit', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify(payload)
                });

                const result = await response.json();

                if (response.status === 409) {
                    showToast(result.message || 'Data duplikat. Gunakan Edit di Riwayat.', 'warning');
                } else if (response.status === 403) {
                    showToast(result.message || 'Lokasi tidak ditugaskan untuk petugas ini.', 'warning');
                } else if (response.status === 400) {
                    showToast(result.message || 'Lokasi tidak valid.', 'warning');
                } else if (response.ok) {
                    showToast('Count saved successfully.', 'success');
                    resetSkuAndCount();
                    maybeFetchHistory(true);
                } else {
                    showToast('Sync failed. Try again.', 'error');
                }
            } catch (err) {
                showToast('Network error. Try again.', 'error');
            } finally {
                btn.dataset.loading = '0';
                label.textContent = 'Kirim ke sheet';
                spinner.classList.add('hidden');
                updateSubmitState();
            }
        }

        async function editItem(logId, currentCount) {
            const newCount = prompt(`Enter new physical count for this row:`, currentCount);
            if (newCount === null || newCount.trim() === "" || isNaN(newCount)) return;
            
            try {
                const response = await fetch('/edit', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ id: logId, count: parseInt(newCount), session_id: activeSession.sessionId })
                });
                if (response.ok) {
                    maybeFetchHistory(true);
                } else {
                    alert('Failed to update record on sheet.');
                }
            } catch (err) {
                alert('Network error, update aborted.');
            }
        }

        async function deleteItem(logId) {
            if (!confirm("Are you sure you want to delete this specific count record from the master sheet?")) return;
            
            try {
                const response = await fetch('/delete', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ id: logId, session_id: activeSession.sessionId })
                });
                if (response.ok) {
                    maybeFetchHistory(true);
                } else {
                    alert('Failed to delete record from sheet.');
                }
            } catch (err) {
                alert('Network error, delete aborted.');
            }
        }
    </script>
</body>
</html>
"""

DASHBOARD_HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Opname Dashboard — Aeris Beaute</title>
    <script src="https://cdn.jsdelivr.net/npm/@tailwindcss/browser@4"></script>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
    <style>
        @keyframes pulse-live { 0%, 100% { opacity: 1; } 50% { opacity: 0.45; } }
        .live-dot { animation: pulse-live 2s ease-in-out infinite; }
        .metric-card { transition: transform 0.2s ease, box-shadow 0.2s ease; }
        .metric-card:hover { transform: translateY(-2px); box-shadow: 0 8px 24px rgb(0 0 0 / 0.08); }
        .progress-ring { transform: rotate(-90deg); transform-origin: 50% 50%; }
    </style>
</head>
<body class="bg-zinc-50 font-sans text-zinc-900 antialiased min-h-screen">
    <header class="sticky top-0 z-40 bg-white/95 backdrop-blur-md border-b border-zinc-200 shadow-sm">
        <div class="max-w-6xl mx-auto px-4 py-3">
            <div class="flex flex-wrap items-center justify-between gap-3">
                <div>
                    <h1 class="text-lg font-bold text-zinc-900 tracking-tight">Opname Dashboard</h1>
                    <p id="dashSubtitle" class="text-xs text-zinc-500">Memuat…</p>
                </div>
                <nav class="flex flex-col items-end gap-1 text-xs font-semibold shrink-0">
                    <div class="flex flex-wrap gap-3 justify-end">
                        <a href="/count" class="text-zinc-500 hover:text-violet-700">Count</a>
                        <span class="text-violet-700">Dashboard</span>
                        <a href="/summary" class="text-zinc-500 hover:text-violet-700">Summary</a>
                    </div>
                    <p id="dashSessionBadge" class="text-[10px] text-zinc-500 max-w-[12rem] truncate"></p>
                    <button type="button" onclick="endSession()" class="text-[10px] text-rose-600 hover:text-rose-800 font-semibold">End Session</button>
                </nav>
            </div>
        </div>
    </header>

    <main class="max-w-6xl mx-auto px-4 py-5 space-y-5">
        <div class="flex flex-wrap items-center justify-between gap-3">
            <div class="flex items-center gap-2 text-xs font-medium text-emerald-700 bg-emerald-50 border border-emerald-200 rounded-full px-3 py-1.5">
                <span class="live-dot inline-block h-2 w-2 rounded-full bg-emerald-500"></span>
                <span id="liveLabel">Live · refresh 20s</span>
            </div>
            <button type="button" onclick="loadDashboard(true)" class="text-sm font-semibold text-violet-700 hover:text-violet-900 px-3 py-1.5 rounded-lg border border-violet-200 bg-violet-50">Refresh now</button>
        </div>

        <div class="grid grid-cols-2 lg:grid-cols-4 gap-3">
            <div class="metric-card col-span-2 lg:col-span-1 bg-white rounded-xl border border-zinc-200 p-4 shadow-sm">
                <p class="text-xs font-semibold uppercase tracking-wide text-zinc-500">Lokasi tercover</p>
                <p class="text-3xl font-bold text-violet-700 mt-1 tabular-nums" id="metricLocPct">—</p>
                <p class="text-sm text-zinc-600 mt-0.5" id="metricLocFrac">— / —</p>
                <div class="mt-3 h-2 rounded-full bg-zinc-100 overflow-hidden">
                    <div id="barLoc" class="h-full bg-violet-500 rounded-full transition-all duration-700" style="width:0%"></div>
                </div>
            </div>
            <div class="metric-card col-span-2 lg:col-span-1 bg-white rounded-xl border border-zinc-200 p-4 shadow-sm">
                <p class="text-xs font-semibold uppercase tracking-wide text-zinc-500">SKU tercover</p>
                <p class="text-3xl font-bold text-indigo-700 mt-1 tabular-nums" id="metricSkuPct">—</p>
                <p class="text-sm text-zinc-600 mt-0.5" id="metricSkuFrac">— / —</p>
                <div class="mt-3 h-2 rounded-full bg-zinc-100 overflow-hidden">
                    <div id="barSku" class="h-full bg-indigo-500 rounded-full transition-all duration-700" style="width:0%"></div>
                </div>
            </div>
            <div class="metric-card bg-white rounded-xl border border-zinc-200 p-4 shadow-sm">
                <p class="text-xs font-semibold uppercase tracking-wide text-zinc-500">Baris entri</p>
                <p class="text-2xl font-bold text-zinc-900 mt-1 tabular-nums" id="metricEntries">—</p>
                <p class="text-xs text-zinc-500 mt-1"><span id="metricRecent">0</span> dalam 15 menit</p>
            </div>
            <div class="metric-card bg-white rounded-xl border border-zinc-200 p-4 shadow-sm">
                <p class="text-xs font-semibold uppercase tracking-wide text-zinc-500">Total qty</p>
                <p class="text-2xl font-bold text-zinc-900 mt-1 tabular-nums" id="metricQty">—</p>
                <p class="text-xs text-zinc-500 mt-1"><span id="metricCounters">0</span> petugas aktif</p>
            </div>
        </div>

        <div class="grid md:grid-cols-2 gap-4">
            <div class="bg-white rounded-xl border border-zinc-200 p-4 shadow-sm">
                <h2 class="text-sm font-semibold text-zinc-800 mb-3">Cakupan lokasi</h2>
                <div class="h-56 flex items-center justify-center"><canvas id="chartLocations"></canvas></div>
            </div>
            <div class="bg-white rounded-xl border border-zinc-200 p-4 shadow-sm">
                <h2 class="text-sm font-semibold text-zinc-800 mb-3">Cakupan SKU</h2>
                <div class="h-56 flex items-center justify-center"><canvas id="chartSkus"></canvas></div>
            </div>
        </div>

        <div class="grid md:grid-cols-2 gap-4">
            <div class="bg-white rounded-xl border border-zinc-200 p-4 shadow-sm">
                <h2 class="text-sm font-semibold text-zinc-800 mb-3">Aktivitas per jam (hari ini)</h2>
                <div class="h-56"><canvas id="chartHourly"></canvas></div>
            </div>
            <div class="bg-white rounded-xl border border-zinc-200 p-4 shadow-sm">
                <h2 class="text-sm font-semibold text-zinc-800 mb-3">Top SKU (jumlah baris)</h2>
                <div class="h-56"><canvas id="chartTopSku"></canvas></div>
            </div>
        </div>
        <p id="dashUpdated" class="text-xs text-zinc-400 text-center pb-6"></p>
    </main>

    <script>
        const SESSION_STORAGE_KEY = 'aeris_opname_session';
        const POLL_MS = 20000;
        let charts = {};
        let pollTimer = null;

        function loadStoredSession() {
            try {
                const raw = localStorage.getItem(SESSION_STORAGE_KEY);
                return raw ? JSON.parse(raw) : null;
            } catch (e) { return null; }
        }

        function endSession() {
            localStorage.removeItem(SESSION_STORAGE_KEY);
            window.location.href = '/';
        }

        const activeSession = loadStoredSession();
        if (!activeSession || !activeSession.sessionId) {
            window.location.replace('/');
        }

        function fmt(n) {
            return Number(n).toLocaleString('id-ID');
        }

        function upsertDoughnut(id, labels, values, colors) {
            const ctx = document.getElementById(id);
            if (!ctx) return;
            if (charts[id]) {
                charts[id].data.labels = labels;
                charts[id].data.datasets[0].data = values;
                charts[id].update('active');
                return;
            }
            charts[id] = new Chart(ctx, {
                type: 'doughnut',
                data: {
                    labels,
                    datasets: [{ data: values, backgroundColor: colors, borderWidth: 0 }],
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    cutout: '62%',
                    plugins: { legend: { position: 'bottom', labels: { boxWidth: 12, font: { size: 11 } } } },
                },
            });
        }

        function upsertBar(id, labels, values, color) {
            const ctx = document.getElementById(id);
            if (!ctx) return;
            if (charts[id]) {
                charts[id].data.labels = labels;
                charts[id].data.datasets[0].data = values;
                charts[id].update('active');
                return;
            }
            charts[id] = new Chart(ctx, {
                type: 'bar',
                data: {
                    labels,
                    datasets: [{ data: values, backgroundColor: color, borderRadius: 6 }],
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: { legend: { display: false } },
                    scales: {
                        x: { ticks: { font: { size: 10 }, maxRotation: 45, minRotation: 0 } },
                        y: { beginAtZero: true, ticks: { precision: 0 } },
                    },
                },
            });
        }

        function upsertLine(id, labels, values) {
            const ctx = document.getElementById(id);
            if (!ctx) return;
            if (charts[id]) {
                charts[id].data.labels = labels;
                charts[id].data.datasets[0].data = values;
                charts[id].update('active');
                return;
            }
            charts[id] = new Chart(ctx, {
                type: 'line',
                data: {
                    labels,
                    datasets: [{
                        data: values,
                        borderColor: '#7c3aed',
                        backgroundColor: 'rgba(124, 58, 237, 0.12)',
                        fill: true,
                        tension: 0.35,
                        pointRadius: 3,
                    }],
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: { legend: { display: false } },
                    scales: {
                        y: { beginAtZero: true, ticks: { precision: 0 } },
                    },
                },
            });
        }

        function renderDashboard(d) {
            document.getElementById('dashSubtitle').textContent = 'Monitoring sesi · ' + (d.session_name || d.session_id);
            document.getElementById('dashSessionBadge').textContent = d.session_name || d.session_id;

            const loc = d.locations || {};
            const sku = d.skus || {};
            document.getElementById('metricLocPct').textContent = loc.pct + '%';
            document.getElementById('metricLocFrac').textContent = fmt(loc.covered) + ' / ' + fmt(loc.total);
            document.getElementById('barLoc').style.width = Math.min(100, loc.pct || 0) + '%';

            document.getElementById('metricSkuPct').textContent = sku.pct + '%';
            document.getElementById('metricSkuFrac').textContent = fmt(sku.covered) + ' / ' + fmt(sku.total);
            document.getElementById('barSku').style.width = Math.min(100, sku.pct || 0) + '%';

            document.getElementById('metricEntries').textContent = fmt(d.entries || 0);
            document.getElementById('metricRecent').textContent = fmt(d.recent_entries || 0);
            document.getElementById('metricQty').textContent = fmt(d.total_qty || 0);
            document.getElementById('metricCounters').textContent = fmt(d.counters_active || 0);

            const locRemain = Math.max(0, (loc.total || 0) - (loc.covered || 0));
            const skuRemain = Math.max(0, (sku.total || 0) - (sku.covered || 0));
            upsertDoughnut('chartLocations', ['Tercover', 'Belum'], [loc.covered || 0, locRemain], ['#8b5cf6', '#e4e4e7']);
            upsertDoughnut('chartSkus', ['Tercover', 'Belum'], [sku.covered || 0, skuRemain], ['#6366f1', '#e4e4e7']);

            const hourly = d.hourly || [];
            upsertLine('chartHourly', hourly.map(h => h.hour), hourly.map(h => h.entries));

            const top = d.top_skus || [];
            upsertBar('chartTopSku', top.map(t => t.sku), top.map(t => t.lines), '#a78bfa');

            document.getElementById('dashUpdated').textContent = 'Diperbarui: ' + (d.updated_at || '—');
        }

        async function loadDashboard(forceRefresh = false) {
            try {
                const qs = new URLSearchParams({ session_id: activeSession.sessionId });
                if (forceRefresh) qs.set('refresh', '1');
                const res = await fetch('/dashboard/data?' + qs.toString());
                const data = await res.json();
                if (!res.ok) throw new Error(data.error || 'load failed');
                renderDashboard(data);
            } catch (e) {
                document.getElementById('dashSubtitle').textContent = 'Gagal memuat dashboard';
            }
        }

        loadDashboard();
        pollTimer = setInterval(() => loadDashboard(false), POLL_MS);
        document.addEventListener('visibilitychange', () => {
            if (document.visibilityState === 'visible') loadDashboard(false);
        });
    </script>
</body>
</html>
"""

SUMMARY_HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>SKU Summary — Aeris Opname</title>
    <script src="https://cdn.jsdelivr.net/npm/@tailwindcss/browser@4"></script>
</head>
<body class="bg-zinc-50 font-sans text-zinc-900 antialiased min-h-screen">

    <header class="sticky top-0 z-40 bg-white/95 backdrop-blur-md border-b border-zinc-200 shadow-sm">
        <div class="max-w-3xl mx-auto px-4 py-3">
            <div class="flex items-center justify-between gap-3">
                <div>
                    <h1 class="text-lg font-bold text-zinc-900 tracking-tight">SKU Summary</h1>
                    <p class="text-xs text-zinc-500">Running totals from Raw Counts</p>
                </div>
                <nav class="flex flex-col items-end gap-1 text-xs font-semibold shrink-0">
                    <div class="flex flex-wrap gap-3 justify-end">
                        <a href="/count" class="text-zinc-500 hover:text-violet-700">Count</a>
                        <a href="/dashboard" class="text-zinc-500 hover:text-violet-700">Dashboard</a>
                        <span class="text-violet-700">Summary</span>
                    </div>
                    <p id="summarySessionBadge" class="text-[10px] text-zinc-500 max-w-[10rem] truncate"></p>
                    <button type="button" onclick="endSession()" class="text-[10px] text-rose-600 hover:text-rose-800 font-semibold">End Session</button>
                </nav>
            </div>
        </div>
    </header>

    <main class="max-w-3xl mx-auto px-4 py-4 space-y-4">
        <div class="flex flex-col sm:flex-row sm:items-center gap-3">
            <input type="search" id="searchSku" placeholder="Cari SKU…" oninput="filterTable()"
                class="flex-1 border border-zinc-200 rounded-lg px-3 py-2.5 text-sm focus:border-violet-500 focus:ring-2 focus:ring-violet-500/20 focus:outline-none">
            <button type="button" onclick="loadSummary(true)" class="shrink-0 px-4 py-2.5 rounded-lg bg-violet-600 hover:bg-violet-700 text-white text-sm font-semibold transition">
                Refresh
            </button>
        </div>

        <div class="bg-white rounded-xl border border-zinc-200 shadow-sm overflow-hidden">
            <div class="overflow-x-auto">
                <table class="w-full text-sm">
                    <thead class="bg-zinc-50 border-b border-zinc-200">
                        <tr>
                            <th class="text-left px-4 py-3 font-semibold text-zinc-700">SKU Code</th>
                            <th class="text-right px-4 py-3 font-semibold text-zinc-700">Total qty</th>
                            <th class="text-right px-4 py-3 font-semibold text-zinc-700 hidden sm:table-cell">Lokasi</th>
                        </tr>
                    </thead>
                    <tbody id="summaryBody">
                        <tr><td colspan="3" class="px-4 py-8 text-center text-zinc-400">Loading…</td></tr>
                    </tbody>
                    <tfoot id="summaryFoot" class="bg-violet-50 border-t border-violet-100 hidden">
                        <tr>
                            <td class="px-4 py-3 font-bold text-zinc-800">Grand total</td>
                            <td id="grandTotal" class="px-4 py-3 text-right font-bold text-violet-800 tabular-nums">0</td>
                            <td id="skuCount" class="px-4 py-3 text-right text-sm text-zinc-600 hidden sm:table-cell"></td>
                        </tr>
                    </tfoot>
                </table>
            </div>
        </div>
        <p id="lastUpdated" class="text-xs text-zinc-400 text-center"></p>
    </main>

    <script>
        const SESSION_STORAGE_KEY = 'aeris_opname_session';
        let summaryRows = [];

        function loadStoredSession() {
            try {
                const raw = localStorage.getItem(SESSION_STORAGE_KEY);
                return raw ? JSON.parse(raw) : null;
            } catch (e) { return null; }
        }

        function endSession() {
            localStorage.removeItem(SESSION_STORAGE_KEY);
            window.location.href = '/';
        }

        const activeSession = loadStoredSession();
        if (!activeSession || !activeSession.sessionId) {
            window.location.replace('/');
        }

        function escapeHtml(str) {
            return String(str)
                .replace(/&/g, '&amp;')
                .replace(/</g, '&lt;')
                .replace(/>/g, '&gt;')
                .replace(/"/g, '&quot;');
        }

        function renderTable(rows, grandTotal) {
            const body = document.getElementById('summaryBody');
            const foot = document.getElementById('summaryFoot');

            if (!rows.length) {
                body.innerHTML = '<tr><td colspan="3" class="px-4 py-8 text-center text-zinc-400">Belum ada data di Raw Counts.</td></tr>';
                foot.classList.add('hidden');
                return;
            }

            body.innerHTML = rows.map(r => `
                <tr class="summary-row border-b border-zinc-100 hover:bg-zinc-50/80" data-sku="${escapeHtml(r.sku).toLowerCase()}">
                    <td class="px-4 py-3 font-mono font-semibold text-violet-700">${escapeHtml(r.sku)}</td>
                    <td class="px-4 py-3 text-right font-bold tabular-nums text-zinc-900">${escapeHtml(String(r.total))}</td>
                    <td class="px-4 py-3 text-right text-zinc-500 hidden sm:table-cell">${escapeHtml(String(r.location_count))}</td>
                </tr>
            `).join('');

            document.getElementById('grandTotal').textContent = grandTotal;
            document.getElementById('skuCount').textContent = rows.length + ' SKU';
            foot.classList.remove('hidden');
        }

        function filterTable() {
            const q = document.getElementById('searchSku').value.trim().toLowerCase();
            const filtered = q
                ? summaryRows.filter(r => r.sku.toLowerCase().includes(q))
                : summaryRows;
            const grand = filtered.reduce((s, r) => s + r.total, 0);
            renderTable(filtered, grand);
        }

        async function loadSummary(forceRefresh = false) {
            const body = document.getElementById('summaryBody');
            body.innerHTML = '<tr><td colspan="3" class="px-4 py-8 text-center text-zinc-400 animate-pulse">Loading…</td></tr>';
            const badge = document.getElementById('summarySessionBadge');
            if (badge) {
                badge.textContent = activeSession.sessionName || activeSession.sessionId;
                badge.title = badge.textContent;
            }

            try {
                const qs = new URLSearchParams({ session_id: activeSession.sessionId });
                if (forceRefresh) qs.set('refresh', '1');
                const res = await fetch('/summary/data?' + qs.toString());
                const data = await res.json();
                summaryRows = data.rows || [];
                renderTable(summaryRows, data.grand_total || 0);
                document.getElementById('lastUpdated').textContent =
                    'Diperbarui: ' + new Date().toLocaleString('id-ID', { timeZone: 'Asia/Jakarta' });
            } catch (e) {
                body.innerHTML = '<tr><td colspan="3" class="px-4 py-8 text-center text-rose-600">Gagal memuat data.</td></tr>';
            }
        }

        window.onload = loadSummary;
    </script>
</body>
</html>
"""

ADMIN_STOCK_HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, viewport-fit=cover">
    <title>Stock Reconcile — Aeris Opname</title>
    <script src="https://cdn.jsdelivr.net/npm/@tailwindcss/browser@4"></script>
</head>
<body class="bg-zinc-50 font-sans text-zinc-900 antialiased min-h-screen">
    <header class="border-b border-zinc-200 bg-white shadow-sm">
        <div class="max-w-4xl mx-auto px-4 py-3 flex items-start justify-between gap-4">
            <div>
                <h1 class="text-lg font-bold text-zinc-900 tracking-tight">Aeris Beaute</h1>
                <p class="text-xs text-zinc-500">Admin — stok sistem &amp; penugasan lokasi</p>
            </div>
            <nav class="flex flex-col items-end gap-1 text-xs font-semibold shrink-0">
                <a href="/" class="text-zinc-500 hover:text-violet-700">Sessions</a>
                <a href="/dashboard" class="text-zinc-500 hover:text-violet-700">Dashboard</a>
            </nav>
        </div>
    </header>
    <main class="max-w-4xl mx-auto px-4 py-8 space-y-8">
        <section class="bg-white rounded-xl border border-zinc-200 shadow-sm p-5 space-y-4">
            <h2 class="text-base font-bold text-zinc-900">Upload stok sistem (Excel)</h2>
            <p class="text-sm text-zinc-600">
                File ERP: nama gudang di <strong>baris 5</strong>, SKU di <strong>kolom C</strong> (Kode Barang) mulai baris 6.
                Hanya SKU yang ada di tab <strong>SKU List</strong> yang dimasukkan; baris seperti <strong>Total Kode Barang</strong> diabaikan.
                Tab baru di Google Sheet akan dibuat dengan nama <strong>Session ID</strong>.
            </p>
            <div class="grid gap-4 sm:grid-cols-2">
                <div>
                    <label for="sessionSelect" class="block text-sm font-semibold text-zinc-700 mb-2">Session</label>
                    <select id="sessionSelect" class="w-full border border-zinc-200 rounded-lg px-3 py-2.5 text-sm bg-white focus:border-violet-500 focus:ring-2 focus:ring-violet-500/20 focus:outline-none">
                        <option value="">Memuat sesi…</option>
                    </select>
                </div>
                <div>
                    <label for="excelFile" class="block text-sm font-semibold text-zinc-700 mb-2">File Excel (.xlsx)</label>
                    <input id="excelFile" type="file" accept=".xlsx,.xlsm"
                        class="w-full border border-zinc-200 rounded-lg px-3 py-2 text-sm file:mr-3 file:py-1 file:px-3 file:rounded-md file:border-0 file:bg-violet-50 file:text-violet-700 file:font-semibold">
                </div>
            </div>
            <button type="button" id="uploadBtn" onclick="uploadStock()"
                class="w-full sm:w-auto px-5 py-2.5 rounded-lg bg-violet-600 hover:bg-violet-700 disabled:opacity-50 text-white text-sm font-semibold">
                Upload &amp; buat tab sesi
            </button>
            <p id="uploadStatus" class="hidden text-sm"></p>
            <ul id="uploadWarnings" class="hidden text-sm text-amber-700 list-disc pl-5 space-y-1"></ul>
            <div id="uploadUnmappedBox" class="hidden mt-3 rounded-lg border border-amber-200 bg-amber-50 px-3 py-3 text-sm text-amber-950"></div>
        </section>

        <section id="assignmentsSection" class="bg-white rounded-xl border border-zinc-200 shadow-sm p-5 space-y-4">
            <div class="flex flex-wrap items-center justify-between gap-3">
                <h2 class="text-base font-bold text-zinc-900">Penugasan lokasi petugas</h2>
                <button type="button" onclick="loadPetugasAssignments(true)" class="text-xs font-semibold text-violet-700 hover:text-violet-900">Muat ulang</button>
            </div>
            <p class="text-sm text-zinc-600">
                Tentukan lokasi yang boleh dihitung setiap petugas per sesi. Sesi yang punya minimal satu baris di sini akan <strong>mengunci</strong> pemilihan lokasi di halaman Count.
                Data disimpan ke tab <strong>PETUGAS ASSIGNMENTS</strong> di Google Sheet.
            </p>
            <div class="grid gap-4 sm:grid-cols-2">
                <div>
                    <label for="assignSessionFilter" class="block text-sm font-semibold text-zinc-700 mb-2">Filter sesi (tampilan)</label>
                    <select id="assignSessionFilter" onchange="renderAssignmentTable()"
                        class="w-full border border-zinc-200 rounded-lg px-3 py-2.5 text-sm bg-white focus:border-violet-500 focus:ring-2 focus:ring-violet-500/20 focus:outline-none">
                        <option value="">Semua sesi</option>
                    </select>
                </div>
                <div class="flex items-end">
                    <button type="button" onclick="addAssignmentRow()"
                        class="w-full sm:w-auto px-4 py-2.5 rounded-lg border border-violet-200 bg-violet-50 text-violet-800 text-sm font-semibold hover:bg-violet-100">
                        + Tambah baris
                    </button>
                </div>
            </div>
            <div id="assignWarningsBox" class="hidden rounded-lg border border-amber-200 bg-amber-50 px-3 py-3 text-sm text-amber-950 space-y-1"></div>
            <div id="assignGapBox" class="hidden rounded-lg border border-amber-200 bg-amber-50 px-3 py-3 text-sm text-amber-950"></div>
            <div class="overflow-x-auto border border-zinc-200 rounded-lg">
                <table class="w-full text-sm min-w-[36rem]">
                    <thead class="bg-zinc-50 border-b border-zinc-200">
                        <tr>
                            <th class="text-left px-3 py-2 font-semibold text-zinc-700">Session ID</th>
                            <th class="text-left px-3 py-2 font-semibold text-zinc-700">Petugas</th>
                            <th class="text-left px-3 py-2 font-semibold text-zinc-700">Location</th>
                            <th class="w-16 px-2 py-2"></th>
                        </tr>
                    </thead>
                    <tbody id="assignmentsBody">
                        <tr><td colspan="4" class="px-3 py-6 text-center text-zinc-400">Memuat…</td></tr>
                    </tbody>
                </table>
            </div>
            <datalist id="assignCounterList"></datalist>
            <datalist id="assignLocationList"></datalist>
            <datalist id="assignSessionList"></datalist>
            <div class="flex flex-wrap items-center gap-3">
                <button type="button" id="saveAssignmentsBtn" onclick="savePetugasAssignments()"
                    class="px-5 py-2.5 rounded-lg bg-violet-600 hover:bg-violet-700 disabled:opacity-50 text-white text-sm font-semibold">
                    Simpan penugasan
                </button>
                <p id="assignSaveStatus" class="hidden text-sm"></p>
            </div>
        </section>

        <section id="setupCheckSection" class="bg-white rounded-xl border border-zinc-200 shadow-sm p-5 space-y-3">
            <div class="flex flex-wrap items-center justify-between gap-3">
                <h2 class="text-base font-bold text-zinc-900">Pemeriksaan mapping lokasi</h2>
                <button type="button" onclick="loadUnmappedLocations()" class="text-xs font-semibold text-violet-700 hover:text-violet-900">Periksa ulang</button>
            </div>
            <p class="text-sm text-zinc-600">
                Mapping lokasi/zone → gudang diatur di tab <strong>GUDANG LOCATIONS</strong> pada Google Sheet (bukan di halaman ini).
            </p>
            <div id="unmappedOkBox" class="hidden rounded-lg border border-emerald-200 bg-emerald-50 px-3 py-2 text-sm text-emerald-900">
                Semua lokasi di tab LOCATIONS sudah terpetakan ke gudang.
            </div>
            <div id="unmappedFlagBox" class="hidden rounded-lg border border-amber-200 bg-amber-50 px-3 py-3 text-sm text-amber-950">
                <p id="unmappedFlagTitle" class="font-semibold mb-2"></p>
                <ul id="unmappedFlagList" class="list-disc pl-5 space-y-0.5 max-h-48 overflow-y-auto"></ul>
            </div>
        </section>
    </main>
    <script>
        async function loadSessions() {
            const sel = document.getElementById('sessionSelect');
            try {
                const res = await fetch('/api/sessions');
                const data = await res.json();
                const sessions = data.sessions || [];
                sel.innerHTML = '';
                if (!sessions.length) {
                    sel.innerHTML = '<option value="">Tidak ada sesi aktif</option>';
                    return;
                }
                sessions.forEach(s => {
                    const opt = document.createElement('option');
                    opt.value = s.id;
                    opt.textContent = s.name + ' (' + s.id + ')';
                    sel.appendChild(opt);
                });
            } catch (e) {
                sel.innerHTML = '<option value="">Gagal memuat sesi</option>';
            }
        }

        function setStatus(elId, message, kind) {
            const el = document.getElementById(elId);
            el.textContent = message;
            el.classList.remove('hidden', 'text-rose-600', 'text-emerald-700', 'text-zinc-600');
            if (kind === 'error') el.classList.add('text-rose-600');
            else if (kind === 'success') el.classList.add('text-emerald-700');
            else el.classList.add('text-zinc-600');
        }

        function showUnmappedLocations(unmapped, target) {
            const list = unmapped || [];
            const okBox = document.getElementById('unmappedOkBox');
            const flagBox = document.getElementById('unmappedFlagBox');
            const title = document.getElementById('unmappedFlagTitle');
            const ul = document.getElementById('unmappedFlagList');
            const uploadBox = document.getElementById('uploadUnmappedBox');
            if (target === 'upload') {
                if (!list.length) {
                    uploadBox.classList.add('hidden');
                    uploadBox.innerHTML = '';
                    return;
                }
                uploadBox.classList.remove('hidden');
                uploadBox.innerHTML = '<p class="font-semibold mb-1">' + list.length +
                    ' lokasi di hitungan sesi ini belum terpetakan ke gudang:</p><p class="text-xs break-words">' +
                    list.join(', ') + '</p>';
                return;
            }
            okBox.classList.add('hidden');
            flagBox.classList.add('hidden');
            ul.innerHTML = '';
            if (!list.length) {
                okBox.classList.remove('hidden');
                return;
            }
            flagBox.classList.remove('hidden');
            title.textContent = list.length + ' lokasi/zone di tab LOCATIONS belum ada di GUDANG LOCATIONS:';
            list.forEach(loc => {
                const li = document.createElement('li');
                li.textContent = loc;
                ul.appendChild(li);
            });
        }

        async function loadUnmappedLocations() {
            try {
                const res = await fetch('/api/admin/gudang-locations?refresh=1');
                const data = await res.json();
                showUnmappedLocations(data.unmapped_locations || [], 'setup');
            } catch (e) {
                const flagBox = document.getElementById('unmappedFlagBox');
                flagBox.classList.remove('hidden');
                document.getElementById('unmappedFlagTitle').textContent = 'Gagal memeriksa mapping lokasi.';
                document.getElementById('unmappedFlagList').innerHTML = '';
            }
        }

        async function uploadStock() {
            const sessionId = document.getElementById('sessionSelect').value;
            const fileInput = document.getElementById('excelFile');
            const btn = document.getElementById('uploadBtn');
            const warnEl = document.getElementById('uploadWarnings');
            warnEl.classList.add('hidden');
            warnEl.innerHTML = '';
            document.getElementById('uploadUnmappedBox').classList.add('hidden');
            if (!sessionId) {
                setStatus('uploadStatus', 'Pilih sesi terlebih dahulu.', 'error');
                return;
            }
            if (!fileInput.files || !fileInput.files[0]) {
                setStatus('uploadStatus', 'Pilih file Excel (.xlsx).', 'error');
                return;
            }
            btn.disabled = true;
            setStatus('uploadStatus', 'Mengunggah dan memproses…', 'info');
            const form = new FormData();
            form.append('session_id', sessionId);
            form.append('file', fileInput.files[0]);
            try {
                const res = await fetch('/api/admin/stock/upload', { method: 'POST', body: form });
                const data = await res.json();
                if (!res.ok) throw new Error(data.message || 'Upload gagal');
                let msg = 'Tab "' + (data.tab_title || sessionId) + '" dibuat/diperbarui dengan ' + data.row_count + ' baris.';
                if (data.excluded_unrecognized) {
                    msg += ' ' + data.excluded_unrecognized + ' baris SKU tidak dikenali diabaikan.';
                }
                setStatus('uploadStatus', msg, 'success');
                showUnmappedLocations(data.unmapped_locations || [], 'upload');
                if (data.warnings && data.warnings.length) {
                    warnEl.classList.remove('hidden');
                    const note = document.createElement('li');
                    note.className = 'font-semibold';
                    note.textContent = 'SKU di bawah tidak dimasukkan ke tab sesi:';
                    warnEl.appendChild(note);
                    data.warnings.slice(0, 50).forEach(w => {
                        const li = document.createElement('li');
                        li.textContent = w;
                        warnEl.appendChild(li);
                    });
                }
            } catch (e) {
                setStatus('uploadStatus', e.message || 'Upload gagal', 'error');
            } finally {
                btn.disabled = false;
            }
        }

        let assignmentRows = [];
        let assignmentMeta = { counters: [], locations: [], sessions: [] };

        function escapeHtml(str) {
            return String(str)
                .replace(/&/g, '&amp;')
                .replace(/</g, '&lt;')
                .replace(/>/g, '&gt;')
                .replace(/"/g, '&quot;');
        }

        function fillDatalist(id, values) {
            const el = document.getElementById(id);
            el.innerHTML = (values || []).map(v => `<option value="${escapeHtml(v)}"></option>`).join('');
        }

        function getAssignmentFilterSession() {
            return document.getElementById('assignSessionFilter').value.trim();
        }

        function visibleAssignmentRows() {
            const filter = getAssignmentFilterSession();
            if (!filter) return assignmentRows;
            return assignmentRows.filter(r => r.session_id === filter);
        }

        function renderAssignmentTable() {
            const body = document.getElementById('assignmentsBody');
            const rows = visibleAssignmentRows();
            if (!rows.length) {
                const filter = getAssignmentFilterSession();
                body.innerHTML = '<tr><td colspan="4" class="px-3 py-6 text-center text-zinc-400">' +
                    (filter ? 'Belum ada penugasan untuk sesi ini.' : 'Belum ada penugasan. Ketuk Tambah baris.') +
                    '</td></tr>';
                return;
            }
            body.innerHTML = rows.map((row, idx) => {
                const globalIdx = assignmentRows.indexOf(row);
                const invalid = row._invalid;
                const rowCls = invalid ? 'bg-rose-50/80' : '';
                return `<tr class="border-b border-zinc-100 ${rowCls}" data-idx="${globalIdx}">
                    <td class="px-2 py-1.5">
                        <input list="assignSessionList" value="${escapeHtml(row.session_id)}"
                            onchange="updateAssignmentField(${globalIdx}, 'session_id', this.value)"
                            class="w-full min-w-[7rem] border border-zinc-200 rounded px-2 py-1.5 text-xs font-mono">
                    </td>
                    <td class="px-2 py-1.5">
                        <input list="assignCounterList" value="${escapeHtml(row.counter)}"
                            onchange="updateAssignmentField(${globalIdx}, 'counter', this.value)"
                            class="w-full min-w-[8rem] border border-zinc-200 rounded px-2 py-1.5 text-xs">
                    </td>
                    <td class="px-2 py-1.5">
                        <input list="assignLocationList" value="${escapeHtml(row.location)}"
                            onchange="updateAssignmentField(${globalIdx}, 'location', this.value)"
                            class="w-full min-w-[8rem] border border-zinc-200 rounded px-2 py-1.5 text-xs font-mono">
                    </td>
                    <td class="px-2 py-1.5 text-center">
                        <button type="button" onclick="removeAssignmentRow(${globalIdx})"
                            class="text-xs font-semibold text-rose-600 hover:text-rose-800">Hapus</button>
                    </td>
                </tr>`;
            }).join('');
        }

        function updateAssignmentField(idx, field, value) {
            if (!assignmentRows[idx]) return;
            assignmentRows[idx][field] = String(value || '').trim();
            assignmentRows[idx]._invalid = false;
        }

        function addAssignmentRow() {
            const sessionId = getAssignmentFilterSession() || (assignmentMeta.sessions[0] && assignmentMeta.sessions[0].id) || '';
            assignmentRows.push({ session_id: sessionId, counter: '', location: '', _invalid: false });
            renderAssignmentTable();
        }

        function removeAssignmentRow(idx) {
            assignmentRows.splice(idx, 1);
            renderAssignmentTable();
        }

        function showAssignWarnings(messages) {
            const box = document.getElementById('assignWarningsBox');
            if (!messages || !messages.length) {
                box.classList.add('hidden');
                box.innerHTML = '';
                return;
            }
            box.classList.remove('hidden');
            box.innerHTML = '<p class="font-semibold">Perhatian:</p>' +
                messages.map(m => `<p>• ${escapeHtml(m)}</p>`).join('');
        }

        function showAssignGaps(countersWithout) {
            const box = document.getElementById('assignGapBox');
            const entries = Object.entries(countersWithout || {});
            if (!entries.length) {
                box.classList.add('hidden');
                box.innerHTML = '';
                return;
            }
            box.classList.remove('hidden');
            let html = '<p class="font-semibold mb-2">Petugas di COUNTERS tanpa penugasan (sesi terkunci):</p><ul class="list-disc pl-5 space-y-2">';
            entries.forEach(([sid, names]) => {
                html += `<li><span class="font-mono text-xs">${escapeHtml(sid)}</span>: ${escapeHtml(names.join(', '))}</li>`;
            });
            html += '</ul>';
            box.innerHTML = html;
        }

        async function loadPetugasAssignments(forceRefresh = false) {
            const body = document.getElementById('assignmentsBody');
            body.innerHTML = '<tr><td colspan="4" class="px-3 py-6 text-center text-zinc-400 animate-pulse">Memuat…</td></tr>';
            try {
                const qs = forceRefresh ? '?refresh=1' : '';
                const res = await fetch('/api/admin/petugas-assignments' + qs);
                const data = await res.json();
                assignmentRows = (data.assignments || []).map(r => ({
                    session_id: r.session_id || '',
                    counter: r.counter || '',
                    location: r.location || '',
                    _invalid: r.valid === false,
                }));
                assignmentMeta = {
                    counters: data.counters || [],
                    locations: data.locations || [],
                    sessions: data.sessions || [],
                };
                fillDatalist('assignCounterList', assignmentMeta.counters);
                fillDatalist('assignLocationList', assignmentMeta.locations);
                fillDatalist('assignSessionList', assignmentMeta.sessions.map(s => s.id));
                const filterSel = document.getElementById('assignSessionFilter');
                const prev = filterSel.value;
                filterSel.innerHTML = '<option value="">Semua sesi</option>';
                assignmentMeta.sessions.forEach(s => {
                    const opt = document.createElement('option');
                    opt.value = s.id;
                    opt.textContent = s.name + ' (' + s.id + ')';
                    filterSel.appendChild(opt);
                });
                if (prev && [...filterSel.options].some(o => o.value === prev)) {
                    filterSel.value = prev;
                }
                showAssignWarnings(data.warnings || []);
                showAssignGaps(data.counters_without_assignments || {});
                renderAssignmentTable();
            } catch (e) {
                body.innerHTML = '<tr><td colspan="4" class="px-3 py-6 text-center text-rose-600">Gagal memuat penugasan.</td></tr>';
            }
        }

        async function savePetugasAssignments() {
            const btn = document.getElementById('saveAssignmentsBtn');
            const statusEl = document.getElementById('assignSaveStatus');
            btn.disabled = true;
            statusEl.classList.remove('hidden', 'text-rose-600', 'text-emerald-700');
            statusEl.textContent = 'Menyimpan…';
            statusEl.classList.add('text-zinc-600');
            const payload = {
                assignments: assignmentRows
                    .filter(r => r.session_id || r.counter || r.location)
                    .map(r => ({
                        session_id: r.session_id,
                        counter: r.counter,
                        location: r.location,
                    })),
            };
            try {
                const res = await fetch('/api/admin/petugas-assignments', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify(payload),
                });
                const data = await res.json();
                if (!res.ok) throw new Error(data.message || 'Simpan gagal');
                let msg = 'Tersimpan: ' + (data.assignment_count || 0) + ' baris penugasan.';
                if (data.warnings && data.warnings.length) {
                    msg += ' ' + data.warnings.length + ' baris diabaikan — lihat peringatan.';
                }
                statusEl.textContent = msg;
                statusEl.classList.remove('text-zinc-600');
                statusEl.classList.add('text-emerald-700');
                showAssignWarnings((data.warnings || []).concat(data.sheet_warnings || []));
                await loadPetugasAssignments(true);
            } catch (e) {
                statusEl.textContent = e.message || 'Simpan gagal';
                statusEl.classList.remove('text-zinc-600');
                statusEl.classList.add('text-rose-600');
            } finally {
                btn.disabled = false;
            }
        }

        window.onload = () => {
            loadSessions();
            loadPetugasAssignments();
            loadUnmappedLocations();
        };
    </script>
</body>
</html>
"""

# --- ROUTES ---

@app.route('/')
def session_page():
    return render_template_string(SESSION_HTML_TEMPLATE)

@app.route('/count')
def count_page():
    location_lookup = {}
    counter_lookup = {}
    lookup_warnings = []
    lookups = {}
    try:
        lookups, _ = build_lookups_payload()
        location_lookup = lookups["location_lookup"]
        counter_lookup = lookups["counter_lookup"]
        lookup_warnings = lookups["lookup_warnings"]
    except Exception:
        lookups = {}
        location_lookup = location_lookup or {}
        counter_lookup = counter_lookup or {}

    boot_data = {
        "location_lookup": location_lookup,
        "counter_lookup": counter_lookup,
        "lookup_warnings": lookup_warnings,
        "assignments": lookups.get("assignments") or {},
        "enforced_session_ids": lookups.get("enforced_session_ids") or [],
    }
    return render_template_string(
        HTML_TEMPLATE,
        valid_locations=sorted(set(location_lookup.values())),
        location_lookup=location_lookup,
        counter_lookup=counter_lookup,
        valid_counters=sorted(set(counter_lookup.values())),
        lookup_warnings=lookup_warnings,
        boot_data=boot_data,
    )

@app.route('/api/sku-codes')
def api_sku_codes():
    try:
        force_refresh = request.args.get("refresh") in ("1", "true", "yes")
        payload, from_cache = build_sku_payload(force_refresh=force_refresh)
        response = jsonify({
            "sku_lookup": payload["sku_lookup"],
            "sku_codes": payload["sku_codes"],
            "warnings": payload["warnings"],
        })
        if from_cache and not force_refresh:
            response.headers["Cache-Control"] = f"private, max-age={_LOOKUPS_CACHE_TTL}"
        return response, 200
    except Exception as e:
        return jsonify({"sku_lookup": {}, "sku_codes": [], "warnings": [], "error": str(e)}), 500

@app.route('/api/sessions')
def api_sessions():
    try:
        force_refresh = request.args.get("refresh") in ("1", "true", "yes")
        sessions, from_cache = build_sessions_payload(force_refresh=force_refresh)
        response = jsonify({"sessions": sessions})
        if from_cache and not force_refresh:
            response.headers["Cache-Control"] = f"private, max-age={_LOOKUPS_CACHE_TTL}"
        return response, 200
    except Exception as e:
        return jsonify({"sessions": [], "error": str(e)}), 500

@app.route('/api/lookups')
def api_lookups():
    try:
        force_refresh = request.args.get("refresh") in ("1", "true", "yes")
        lookups, from_cache = build_lookups_payload(force_refresh=force_refresh)
        response = jsonify({
            "locations": lookups["location_lookup"],
            "counters": lookups["counter_lookup"],
            "location_count": len(lookups["location_lookup"]),
            "counter_count": len(lookups["counter_lookup"]),
            "warnings": lookups["lookup_warnings"],
            "assignments": lookups.get("assignments") or {},
            "enforced_session_ids": lookups.get("enforced_session_ids") or [],
        })
        if from_cache and not force_refresh:
            response.headers["Cache-Control"] = f"private, max-age={_LOOKUPS_CACHE_TTL}"
        return response, 200
    except Exception as e:
        return jsonify({
            "locations": {},
            "counters": {},
            "location_count": 0,
            "counter_count": 0,
            "warnings": [],
            "assignments": {},
            "enforced_session_ids": [],
            "error": str(e),
        }), 500

@app.route('/dashboard')
def dashboard_page():
    return render_template_string(DASHBOARD_HTML_TEMPLATE)

@app.route('/dashboard/data')
def dashboard_data():
    try:
        session_id = request.args.get("session_id", "").strip()
        if not session_id:
            return jsonify({"error": "session_id required"}), 400
        if not require_valid_session_id(session_id):
            return jsonify({"error": "Invalid session"}), 400
        force_refresh = request.args.get("refresh") in ("1", "true", "yes")
        payload, from_cache = build_dashboard_payload(session_id, force_refresh=force_refresh)
        response = jsonify(payload)
        if from_cache and not force_refresh:
            response.headers["Cache-Control"] = f"private, max-age={_DASHBOARD_CACHE_TTL}"
        return response, 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/summary')
def summary_page():
    return render_template_string(SUMMARY_HTML_TEMPLATE)

@app.route('/admin/stock')
def admin_stock_page():
    return render_template_string(ADMIN_STOCK_HTML_TEMPLATE)

@app.route('/api/admin/stock/upload', methods=['POST'])
def api_admin_stock_upload():
    try:
        session_id = str(request.form.get("session_id", "")).strip()
        if not session_id:
            return jsonify({"status": "error", "message": "Pilih sesi."}), 400
        upload = request.files.get("file")
        if not upload or not upload.filename:
            return jsonify({"status": "error", "message": "File Excel wajib diunggah."}), 400
        if not upload.filename.lower().endswith((".xlsx", ".xlsm")):
            return jsonify({"status": "error", "message": "Format file harus .xlsx atau .xlsm."}), 400
        result = import_system_stock_for_session(session_id, upload.read())
        return jsonify({"status": "success", **result}), 200
    except ValueError as e:
        return jsonify({"status": "error", "message": str(e)}), 400
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/admin/gudang-locations', methods=['GET'])
def api_admin_gudang_locations_get():
    try:
        force_refresh = request.args.get("refresh") in ("1", "true", "yes")
        payload, from_cache = build_gudang_locations_payload(force_refresh=force_refresh)
        response = jsonify(payload)
        if from_cache and not force_refresh:
            response.headers["Cache-Control"] = f"private, max-age={_LOOKUPS_CACHE_TTL}"
        return response, 200
    except Exception as e:
        return jsonify({"mappings": [], "unmapped_locations": [], "error": str(e)}), 500

@app.route('/api/admin/gudang-locations', methods=['POST'])
def api_admin_gudang_locations_post():
    try:
        data = request.get_json(silent=True) or {}
        mappings = data.get("mappings") or []
        if not isinstance(mappings, list):
            return jsonify({"status": "error", "message": "mappings harus berupa array."}), 400
        count = save_gudang_location_mappings(mappings)
        return jsonify({"status": "success", "mapping_count": count}), 200
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/admin/petugas-assignments', methods=['GET'])
def api_admin_petugas_assignments_get():
    try:
        force_refresh = request.args.get("refresh") in ("1", "true", "yes")
        payload, from_cache = build_petugas_assignments_admin_payload(force_refresh=force_refresh)
        response = jsonify(payload)
        if from_cache and not force_refresh:
            response.headers["Cache-Control"] = f"private, max-age={_LOOKUPS_CACHE_TTL}"
        return response, 200
    except Exception as e:
        return jsonify({
            "assignments": [],
            "sessions": [],
            "counters": [],
            "locations": [],
            "enforced_session_ids": [],
            "counters_without_assignments": {},
            "warnings": [],
            "error": str(e),
        }), 500

@app.route('/api/admin/petugas-assignments', methods=['POST'])
def api_admin_petugas_assignments_post():
    try:
        data = request.get_json(silent=True) or {}
        assignments = data.get("assignments")
        if assignments is None:
            return jsonify({"status": "error", "message": "assignments wajib diisi."}), 400
        if not isinstance(assignments, list):
            return jsonify({"status": "error", "message": "assignments harus berupa array."}), 400
        count, warnings = save_petugas_assignments(assignments)
        payload, _ = build_petugas_assignments_admin_payload(force_refresh=True)
        return jsonify({
            "status": "success",
            "assignment_count": count,
            "warnings": warnings,
            "sheet_warnings": payload.get("warnings") or [],
            "counters_without_assignments": payload.get("counters_without_assignments") or {},
        }), 200
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/summary/data')
def summary_data():
    try:
        session_id = request.args.get("session_id", "").strip()
        if not session_id:
            return jsonify({"rows": [], "grand_total": 0, "sku_count": 0, "error": "session_id required"}), 400
        if not require_valid_session_id(session_id):
            return jsonify({"rows": [], "grand_total": 0, "sku_count": 0, "error": "Invalid session"}), 400
        force_refresh = request.args.get("refresh") in ("1", "true", "yes")
        payload, from_cache = build_summary_payload(session_id, force_refresh=force_refresh)
        response = jsonify(payload)
        if from_cache and not force_refresh:
            response.headers["Cache-Control"] = f"private, max-age={_SUMMARY_CACHE_TTL}"
        return response, 200
    except Exception as e:
        return jsonify({"rows": [], "grand_total": 0, "sku_count": 0, "error": str(e)}), 500

@app.route('/history', methods=['GET'])
def history():
    try:
        session_id = request.args.get("session_id", "").strip()
        raw_name = request.args.get('name', '').strip()
        if not session_id or not raw_name:
            return jsonify({"items": [], "truncated": False}), 200
        if not require_valid_session_id(session_id):
            return jsonify({"items": [], "truncated": False, "error": "Invalid session"}), 400

        force_refresh = request.args.get("refresh") in ("1", "true", "yes")
        full_scan = request.args.get("full") in ("1", "true", "yes")
        lookups, _ = build_lookups_payload()
        counter_lookup = lookups["counter_lookup"]
        target_name = resolve_counter(raw_name, counter_lookup) or raw_name
        wb = get_spreadsheet_cached()
        sheet = wb.worksheet("Raw Counts")
        items, truncated, from_cache = fetch_counter_history(
            sheet,
            session_id,
            target_name,
            counter_lookup,
            force_refresh=force_refresh,
            full_scan=full_scan,
        )
        response = jsonify({"items": items, "truncated": truncated})
        if from_cache and not force_refresh:
            response.headers["Cache-Control"] = f"private, max-age={_HISTORY_CACHE_TTL}"
        return response, 200
    except Exception:
        return jsonify({"items": [], "truncated": False}), 500

@app.route('/submit', methods=['POST'])
def submit():
    try:
        data = request.json
        session_id = str(data.get("session_id", "")).strip()
        if not session_id:
            return jsonify({"status": "error", "message": "Sesi tidak dipilih. Kembali ke halaman Start Session."}), 400
        if not require_valid_session_id(session_id):
            return jsonify({"status": "error", "message": "Sesi tidak valid. Pilih sesi dari daftar."}), 400

        wb = get_spreadsheet_cached()
        sheet = wb.worksheet("Raw Counts")

        lookups, _ = build_lookups_payload()
        counter_lookup = lookups["counter_lookup"]
        location_lookup = lookups["location_lookup"]
        counter_name = resolve_counter(data.get('counter_name', ''), counter_lookup)

        if not counter_lookup:
            return jsonify({
                "status": "error",
                "message": "Daftar petugas tidak tersedia. Ketik nama sesuai dengan ID badge.",
            }), 400

        if not counter_name:
            return jsonify({
                "status": "invalid_counter",
                "message": "Nama petugas tidak valid. Scan ID badge atau ketik nama sesuai dengan ID badge.",
            }), 400

        loc_string = resolve_location(data.get('location', ''), location_lookup)

        if not location_lookup:
            return jsonify({
                "status": "error",
                "message": "Daftar lokasi tidak tersedia. Periksa tab LOCATIONS di sheet.",
            }), 400

        if not loc_string:
            raw = str(data.get('location', '')).strip()
            return jsonify({
                "status": "invalid_location",
                "message": f"Lokasi tidak valid: {raw}. Scan QR lokasi yang benar.",
            }), 400

        ok, assignment_msg = validate_petugas_location_assignment(
            session_id,
            counter_name,
            loc_string,
            lookups.get("assignments") or {},
            lookups.get("enforced_session_ids") or [],
            counter_lookup,
        )
        if not ok:
            return jsonify({
                "status": "location_not_assigned",
                "message": assignment_msg,
            }), 403

        sku_lookup = load_sku_lookup_cached(wb)
        if not sku_lookup:
            return jsonify({
                "status": "error",
                "message": "Daftar SKU tidak tersedia. Periksa tab SKU List di sheet.",
            }), 400

        sku_code = resolve_sku(data.get('sku', ''), sku_lookup)
        if not sku_code:
            raw_sku = str(data.get('sku', '')).strip()
            return jsonify({
                "status": "invalid_sku",
                "message": f"SKU tidak valid: {raw_sku}. Pilih atau scan SKU dari daftar.",
            }), 400

        dup_count = find_duplicate_count(sheet, session_id, counter_name, loc_string, sku_code)
        if dup_count is not None:
            return jsonify({
                "status": "duplicate",
                "message": (
                    f"Sudah tercatat: {sku_code} di {loc_string} "
                    f"(jumlah {dup_count}). "
                    f"Ubah lewat Edit di tab Riwayat."
                ),
            }), 409

        # Process standard row generation if duplicate test passes
        log_id = str(uuid.uuid4()) 
        parts = loc_string.split('-')
        
        zone = parts[0] if len(parts) > 0 else loc_string[:1]
        rack = parts[1] if len(parts) > 1 else ""
        shelf = parts[2] if len(parts) > 2 else ""
        
        jakarta_tz = pytz.timezone('Asia/Jakarta')
        now_wib = datetime.now(jakarta_tz)
        timestamp = now_wib.strftime("%d/%m/%Y %H:%M:%S")
        notes = str(data.get('notes', '')).strip()
        
        try:
            physical_count = int(data['count'])
        except (ValueError, TypeError):
            return jsonify({"status": "error", "message": "Jumlah tidak valid."}), 400
        if physical_count < 0:
            return jsonify({"status": "error", "message": "Jumlah tidak boleh negatif."}), 400
        
        row_to_append = [
            log_id,              # Column A: Log ID
            counter_name,        # Column B: Counter Name
            zone,                # Column C: Zone
            rack,                # Column D: Rack
            shelf,               # Column E: Shelf
            loc_string,          # Column F: Precise Location
            sku_code,            # Column G: SKU Code
            physical_count,    # Column H: Physical Count
            timestamp,           # Column I: Timestamp
            notes,               # Column J: Notes
            session_id,          # Column K: Session ID
        ]
        
        sheet.append_row(row_to_append)
        invalidate_history_cache(counter_name, session_id)
        invalidate_summary_cache(session_id)
        update_dup_index_entry(session_id, counter_name, loc_string, sku_code, physical_count)
        maybe_refresh_session_stock(session_id)
        return jsonify({"status": "success"}), 200
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/edit', methods=['POST'])
def edit():
    try:
        data = request.json
        log_id = data['id']
        new_count = data['count']
        session_id = str(data.get("session_id", "")).strip() or None
        
        wb = get_spreadsheet_cached()
        sheet = wb.worksheet("Raw Counts")
        cell = sheet.find(log_id)
        
        if cell:
            sheet.update_cell(cell.row, 8, new_count)  # Column H: Physical Count
            invalidate_count_caches(session_id=session_id, invalidate_dup=True)
            maybe_refresh_session_stock(session_id)
            return jsonify({"status": "success"}), 200
        return jsonify({"status": "error", "message": "Record row not found"}), 404
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/delete', methods=['POST'])
def delete():
    try:
        data = request.json
        log_id = data['id']
        session_id = str(data.get("session_id", "")).strip()
        
        wb = get_spreadsheet_cached()
        sheet = wb.worksheet("Raw Counts")
        cell = sheet.find(log_id)
        
        if cell:
            sheet.delete_rows(cell.row)
            invalidate_count_caches(session_id=session_id or None, invalidate_dup=True)
            maybe_refresh_session_stock(session_id)
            return jsonify({"status": "success"}), 200
        return jsonify({"status": "error", "message": "Record row not found"}), 404
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/favicon.ico')
@app.route('/favicon.png')
def favicon():
    return '', 204

app = app